"""集計値のみを埋め込む静的版Excel生成（PII非搭載・Git安全）

build_report.py が生成する『データシート付き完全版』とは別に、
集計結果だけを静的値で書き出したxlsxを生成する。
リポジトリへコミットするのはこちらのファイルを推奨。

生成内容: build_report.pyと同じ分析シート構成（概要・全期間・年度別）
差分: 「データ」シートは含まず、全セルは数式ではなく計算済みの値
"""
import csv
from datetime import datetime
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# build_report.py の設定・スタイル・選択肢マスタを共有
from build_report import (
    CSV_PATH, STORE_NAME, AGE_BUCKETS, GENDERS,
    CHANNELS, CHANNEL_CODE_MAP,
    DECISION_OPTIONS, PACE_OPTIONS, BUDGET_OPTIONS, LIFESTYLE_OPTIONS,
    parse_date, normalize_channel, load_rows,
    style_section, style_col_header, style_total, style_cell,
)

OUT_PATH = 'customer_analysis_store168_aggregates.xlsx'


def age_at(dob, ref):
    a = ref.year - dob.year
    if (ref.month, ref.day) < (dob.month, dob.day):
        a -= 1
    return a


def age_bucket(age):
    if age is None or age < 10:
        return None
    for label, lo, hi in AGE_BUCKETS:
        if lo <= age <= hi:
            return label
    return None


def gender_label(code):
    return {'1': '男性', '2': '女性'}.get(code, '未設定')


def enrich(records):
    """派生列(性別ラベル/年齢/年齢帯/チャネルセット)を追加"""
    for r in records:
        r['gender'] = gender_label(r['gender_code'])
        if r['dob'] and r['reg']:
            r['age'] = age_at(r['dob'], r['reg'])
            r['age_bucket'] = age_bucket(r['age'])
        else:
            r['age'] = None
            r['age_bucket'] = None
        r['channels'] = set(p.strip() for p in r['channel_norm'].split(',') if p.strip())
    return records


def filt(records, year=None):
    if year is None:
        return records
    return [r for r in records if r['reg'] and r['reg'].year == year]


def pct(n, d):
    return round(n / d * 100, 1) if d else 0.0


def write_gender_block(ws, start_row, records):
    ws.cell(row=start_row, column=1, value='■ 男女比率')
    style_section(ws.cell(row=start_row, column=1))
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
    for i, h in enumerate(['区分', '人数', '構成比(%)', '有効母数比(%)*'], 1):
        style_col_header(ws.cell(row=start_row + 1, column=i, value=h))

    c = Counter(r['gender'] for r in records)
    total = sum(c.values())
    valid = c['男性'] + c['女性']
    r = start_row + 2
    for g in ['男性', '女性', '未設定']:
        ws.cell(row=r, column=1, value=g)
        ws.cell(row=r, column=2, value=c[g])
        ws.cell(row=r, column=3, value=pct(c[g], total))
        ws.cell(row=r, column=4, value=pct(c[g], valid) if g != '未設定' else '-')
        for cc in range(1, 5):
            style_cell(ws.cell(row=r, column=cc))
        r += 1
    ws.cell(row=r, column=1, value='合計')
    ws.cell(row=r, column=2, value=total)
    ws.cell(row=r, column=3, value=100.0 if total else 0)
    ws.cell(row=r, column=4, value=100.0 if valid else 0)
    for cc in range(1, 5):
        style_total(ws.cell(row=r, column=cc))
    ws.cell(row=r + 1, column=1, value='* 有効母数比 = 男女合計を母数とした比率（未設定を除く）')
    ws.cell(row=r + 1, column=1).font = Font(italic=True, size=9, color='7F7F7F')
    return r + 3


def write_age_block(ws, start_row, records):
    ws.cell(row=start_row, column=1, value='■ 男女別 年齢構成（登録時点の満年齢）')
    style_section(ws.cell(row=start_row, column=1))
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
    for i, h in enumerate(['年齢帯', '男性', '男性 %', '女性', '女性 %', '未設定', '合計', '合計 %'], 1):
        style_col_header(ws.cell(row=start_row + 1, column=i, value=h))

    mat = {g: Counter() for g in GENDERS}
    unk = Counter()
    for r in records:
        if r['age_bucket']:
            mat[r['gender']][r['age_bucket']] += 1
        else:
            unk[r['gender']] += 1

    total_m = sum(mat['男性'].values())
    total_f = sum(mat['女性'].values())
    total_u = sum(mat['未設定'].values())
    grand = total_m + total_f + total_u

    row = start_row + 2
    for lbl, _, _ in AGE_BUCKETS:
        m, f, u = mat['男性'][lbl], mat['女性'][lbl], mat['未設定'][lbl]
        ws.cell(row=row, column=1, value=lbl)
        ws.cell(row=row, column=2, value=m)
        ws.cell(row=row, column=3, value=pct(m, total_m))
        ws.cell(row=row, column=4, value=f)
        ws.cell(row=row, column=5, value=pct(f, total_f))
        ws.cell(row=row, column=6, value=u)
        ws.cell(row=row, column=7, value=m + f + u)
        ws.cell(row=row, column=8, value=pct(m + f + u, grand))
        for c in range(1, 9):
            style_cell(ws.cell(row=row, column=c))
        row += 1

    ws.cell(row=row, column=1, value='生年月日不明')
    ws.cell(row=row, column=2, value=unk['男性'])
    ws.cell(row=row, column=3, value='-')
    ws.cell(row=row, column=4, value=unk['女性'])
    ws.cell(row=row, column=5, value='-')
    ws.cell(row=row, column=6, value=unk['未設定'])
    ws.cell(row=row, column=7, value=sum(unk.values()))
    ws.cell(row=row, column=8, value='-')
    for c in range(1, 9):
        style_cell(ws.cell(row=row, column=c))
    row += 1

    ws.cell(row=row, column=1, value='年齢判明合計')
    ws.cell(row=row, column=2, value=total_m)
    ws.cell(row=row, column=3, value=100.0 if total_m else 0)
    ws.cell(row=row, column=4, value=total_f)
    ws.cell(row=row, column=5, value=100.0 if total_f else 0)
    ws.cell(row=row, column=6, value=total_u)
    ws.cell(row=row, column=7, value=grand)
    ws.cell(row=row, column=8, value=100.0 if grand else 0)
    for c in range(1, 9):
        style_total(ws.cell(row=row, column=c))
    return row + 2


def write_channel_block(ws, start_row, records):
    ws.cell(row=start_row, column=1, value='■ 当店を知ったきっかけ（性別別・複数選択）')
    style_section(ws.cell(row=start_row, column=1))
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)
    for i, h in enumerate(['項目', '男性 件数', '男性 %', '女性 件数', '女性 %', '未設定 件数', '未設定 %', '合計 件数', '合計 %'], 1):
        style_col_header(ws.cell(row=start_row + 1, column=i, value=h))

    by_g = {g: Counter() for g in GENDERS}
    respondents = {g: 0 for g in GENDERS}
    for r in records:
        if not r['channel_norm']:
            continue
        respondents[r['gender']] += 1
        for ch in r['channels']:
            by_g[r['gender']][ch] += 1

    denoms = {g: respondents[g] for g in GENDERS}
    denoms_total = sum(denoms.values())

    row = start_row + 2
    for ch in CHANNELS:
        m = by_g['男性'][ch]
        f = by_g['女性'][ch]
        u = by_g['未設定'][ch]
        ws.cell(row=row, column=1, value=ch)
        ws.cell(row=row, column=2, value=m)
        ws.cell(row=row, column=3, value=pct(m, denoms['男性']))
        ws.cell(row=row, column=4, value=f)
        ws.cell(row=row, column=5, value=pct(f, denoms['女性']))
        ws.cell(row=row, column=6, value=u)
        ws.cell(row=row, column=7, value=pct(u, denoms['未設定']))
        ws.cell(row=row, column=8, value=m + f + u)
        ws.cell(row=row, column=9, value=pct(m + f + u, denoms_total))
        for c in range(1, 10):
            style_cell(ws.cell(row=row, column=c))
        row += 1

    ws.cell(row=row, column=1, value='回答者数（%分母）')
    ws.cell(row=row, column=2, value=denoms['男性'])
    ws.cell(row=row, column=3, value=100.0 if denoms['男性'] else 0)
    ws.cell(row=row, column=4, value=denoms['女性'])
    ws.cell(row=row, column=5, value=100.0 if denoms['女性'] else 0)
    ws.cell(row=row, column=6, value=denoms['未設定'])
    ws.cell(row=row, column=7, value=100.0 if denoms['未設定'] else 0)
    ws.cell(row=row, column=8, value=denoms_total)
    ws.cell(row=row, column=9, value=100.0 if denoms_total else 0)
    for c in range(1, 10):
        style_total(ws.cell(row=row, column=c))
    row += 1
    ws.cell(row=row, column=1, value='※ 複数選択項目。%は各性別の「回答者数」を分母')
    ws.cell(row=row, column=1).font = Font(italic=True, size=9, color='7F7F7F')
    return row + 2


def write_single_choice_block(ws, start_row, records, field, options, title):
    ws.cell(row=start_row, column=1, value=title)
    style_section(ws.cell(row=start_row, column=1))
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)
    for i, h in enumerate(['項目', '男性 件数', '男性 %', '女性 件数', '女性 %', '未設定 件数', '未設定 %', '合計 件数', '合計 %'], 1):
        style_col_header(ws.cell(row=start_row + 1, column=i, value=h))

    by_g = {g: Counter() for g in GENDERS}
    respondents = {g: 0 for g in GENDERS}
    for r in records:
        v = r[field]
        if not v:
            continue
        respondents[r['gender']] += 1
        by_g[r['gender']][v] += 1

    denoms = respondents
    total = sum(denoms.values())

    row = start_row + 2
    for opt in options:
        m, f, u = by_g['男性'][opt], by_g['女性'][opt], by_g['未設定'][opt]
        ws.cell(row=row, column=1, value=opt)
        ws.cell(row=row, column=2, value=m)
        ws.cell(row=row, column=3, value=pct(m, denoms['男性']))
        ws.cell(row=row, column=4, value=f)
        ws.cell(row=row, column=5, value=pct(f, denoms['女性']))
        ws.cell(row=row, column=6, value=u)
        ws.cell(row=row, column=7, value=pct(u, denoms['未設定']))
        ws.cell(row=row, column=8, value=m + f + u)
        ws.cell(row=row, column=9, value=pct(m + f + u, total))
        for c in range(1, 10):
            style_cell(ws.cell(row=row, column=c))
        row += 1

    ws.cell(row=row, column=1, value='回答合計')
    ws.cell(row=row, column=2, value=denoms['男性'])
    ws.cell(row=row, column=3, value=100.0 if denoms['男性'] else 0)
    ws.cell(row=row, column=4, value=denoms['女性'])
    ws.cell(row=row, column=5, value=100.0 if denoms['女性'] else 0)
    ws.cell(row=row, column=6, value=denoms['未設定'])
    ws.cell(row=row, column=7, value=100.0 if denoms['未設定'] else 0)
    ws.cell(row=row, column=8, value=total)
    ws.cell(row=row, column=9, value=100.0 if total else 0)
    for c in range(1, 10):
        style_total(ws.cell(row=row, column=c))
    return row + 2


def make_sheet(wb, sheet_name, records, period_label):
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions['A'].width = 34
    for col in 'BCDEFGHIJK':
        ws.column_dimensions[col].width = 14

    ws.cell(row=1, column=1, value=f'{STORE_NAME}  |  期間: {period_label}  |  対象件数: {len(records):,}')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    r = 3
    r = write_gender_block(ws, r, records); r += 1
    r = write_age_block(ws, r, records); r += 1
    r = write_channel_block(ws, r, records); r += 1
    r = write_single_choice_block(ws, r, records, 'decision', DECISION_OPTIONS, '■ ご来店の一番の決め手（性別別）'); r += 1
    r = write_single_choice_block(ws, r, records, 'pace', PACE_OPTIONS, '■ 今後の理想的なご来店ペース（性別別）'); r += 1
    r = write_single_choice_block(ws, r, records, 'budget', BUDGET_OPTIONS, '■ 本日のご予算（性別別）'); r += 1
    r = write_single_choice_block(ws, r, records, 'lifestyle', LIFESTYLE_OPTIONS, '■ 生活スタイル（性別別）')
    return ws


def write_overview(wb, records, years):
    ws = wb.create_sheet('概要・データ状況', 0)
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 60
    ws.cell(row=1, column=1, value=f'{STORE_NAME}  顧客分析レポート（集計値のみ・静的版）')
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws.cell(row=2, column=1, value='ソース: customers_168_YYYYMMDD.csv  /  データ更新は build_report.py で完全版を再生成してください')
    ws.cell(row=2, column=1).font = Font(size=10, color='7F7F7F')

    ws.cell(row=4, column=1, value='■ このファイルの位置づけ')
    style_section(ws.cell(row=4, column=1))
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
    r = 5
    for k, v in [
        ('用途', 'PIIを含まないスナップショット。Gitコミット・共有・レビュー用。'),
        ('集計値', '静的な計算済み値（数式なし）。編集しても再集計されません。'),
        ('データ更新', 'build_report.py を実行して完全版（データシート付き）を再生成してください。'),
    ]:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical='center')
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='■ できないこと（このCSVでは不足）')
    style_section(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    for k, v in [
        ('3回以上／5回以上来店の分析', 'CSVに「来店回数」列がない。予約履歴／施術履歴のエクスポートが別途必要。'),
        ('全店舗合計', 'このファイルは店舗168単体分。他店舗CSVが揃えば合算可能。'),
    ]:
        c = ws.cell(row=r, column=1, value='×')
        c.font = Font(color='C62828', bold=True)
        c.alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2, value=k).font = Font(bold=True)
        ws.cell(row=r, column=3, value=v)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='■ シート・件数')
    style_section(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value='全期間').font = Font(bold=True)
    ws.cell(row=r, column=2, value=f'{len(records):,} 件')
    r += 1
    for y in years:
        cnt = sum(1 for x in records if x['reg'] and x['reg'].year == y)
        ws.cell(row=r, column=1, value=f'{y}年度').font = Font(bold=True)
        ws.cell(row=r, column=2, value=f'{cnt:,} 件')
        r += 1


def main():
    records = enrich(load_rows())
    years = sorted({r['reg'].year for r in records if r['reg']})
    print(f'Total: {len(records):,}  Years: {years}')

    wb = Workbook()
    wb.remove(wb.active)
    write_overview(wb, records, years)
    make_sheet(wb, '全期間', records, '全期間')
    for y in years:
        make_sheet(wb, f'{y}年度', filt(records, y), f'{y}年度')

    wb.save(OUT_PATH)
    print(f'Saved: {OUT_PATH}')


if __name__ == '__main__':
    main()
