"""顧客CSV → Excel分析レポート（Excel Table + 数式ベース）

設計思想:
  - 「データ」シート = Excel Table (tblCustomers)。CSVから読み込んだ生データ + 派生列(数式)
  - 分析シートは全てCOUNTIFS/SUMIFSでtblCustomersを参照
  - 「データ」シートの末尾に新規行を貼り付けるとTableが自動拡張し、
    分析シートの数値がすべて自動更新される（ピボットテーブル相当の挙動）
"""
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

CSV_PATH = 'customers_utf8.csv'
OUT_PATH = 'customer_analysis_store168.xlsx'
STORE_ID = '168'
STORE_NAME = f'店舗{STORE_ID}'
TABLE_NAME = 'tblCustomers'

# 年齢バケット定義
AGE_BUCKETS = [
    ('10〜19', 10, 19),
    ('20〜22', 20, 22),
    ('23〜26', 23, 26),
    ('27〜29', 27, 29),
    ('30〜34', 30, 34),
    ('35〜39', 35, 39),
    ('40〜49', 40, 49),
    ('50〜59', 50, 59),
    ('60〜',   60, 200),
]

GENDERS = ['男性', '女性', '未設定']

# 集計対象の「知ったきっかけ」チャネル（マスタ順）
CHANNELS = [
    'ホットペッパー',
    'minimo',
    '口コミ',
    '紹介',
    'その他',
    '楽天ビューティ',
    'スタッフ個人のInstagram',
    'officialのInstagram',
    'スタッフ個人のTikTok',
    'Googleマップ',
]

# 生データ内のコード → 表示ラベル
CHANNEL_CODE_MAP = {
    '0_le30n6ak': 'ホットペッパー',
    '1_le30ogin': 'minimo',
    '2_le30oih3': '口コミ',
    '3_le30onge': '紹介',
    '4_le30tm6n': 'その他',
}


def parse_date(s):
    if not s:
        return None
    s = s.strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def normalize_channel(raw: str) -> str:
    """コード表記のチャネル値を人間可読ラベルに変換（","区切り複数選択対応）"""
    if not raw:
        return ''
    parts = [p.strip() for p in raw.split(',') if p.strip()]
    return ','.join(CHANNEL_CODE_MAP.get(p, p) for p in parts)


# ============ CSV 読み込み ============

def load_rows():
    with open(CSV_PATH, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader)  # header
        raw = list(reader)

    records = []
    for r in raw:
        rec = {
            'store': STORE_ID,
            'id': r[0],
            'name': r[1],
            'gender_code': r[6],
            'dob': parse_date(r[5]),
            'reg': parse_date(r[10]),
            'channel_raw': r[13].strip(),
            'channel_norm': normalize_channel(r[13]),
            'decision': r[14].strip(),
            'pace': r[15].strip(),
            'budget': r[16].strip(),
            'lifestyle': r[17].strip(),
        }
        records.append(rec)
    return records


# ============ スタイル ============

HEADER_FILL = PatternFill('solid', fgColor='305496')
SECTION_FILL = PatternFill('solid', fgColor='B4C7E7')
TOTAL_FILL = PatternFill('solid', fgColor='FFF2CC')
DATA_HEADER_FILL = PatternFill('solid', fgColor='4472C4')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_section(cell):
    cell.font = Font(bold=True)
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = BORDER


def style_col_header(cell):
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER


def style_total(cell):
    cell.fill = TOTAL_FILL
    cell.font = Font(bold=True)
    cell.border = BORDER


def style_cell(cell):
    cell.border = BORDER
    cell.alignment = Alignment(horizontal='center', vertical='center')


# ============ 「データ」シート ============

# データシート列定義
DATA_COLUMNS = [
    ('店舗', 'value'),
    ('顧客ID', 'value'),
    ('氏名', 'value'),
    ('性別コード', 'value'),
    ('性別', 'formula_gender'),
    ('生年月日', 'value'),
    ('登録日時', 'value'),
    ('登録年度', 'formula_reg_year'),
    ('年齢', 'formula_age'),
    ('年齢帯', 'formula_age_bucket'),
    ('きっかけ_元データ', 'value'),
    ('きっかけ', 'value'),  # 事前正規化済み
    ('フラグ_ホットペッパー', 'formula_channel_flag'),
    ('フラグ_minimo', 'formula_channel_flag'),
    ('フラグ_口コミ', 'formula_channel_flag'),
    ('フラグ_紹介', 'formula_channel_flag'),
    ('フラグ_その他', 'formula_channel_flag'),
    ('フラグ_楽天ビューティ', 'formula_channel_flag'),
    ('フラグ_スタッフ個人のInstagram', 'formula_channel_flag'),
    ('フラグ_officialのInstagram', 'formula_channel_flag'),
    ('フラグ_スタッフ個人のTikTok', 'formula_channel_flag'),
    ('フラグ_Googleマップ', 'formula_channel_flag'),
    ('決め手', 'value'),
    ('ペース', 'value'),
    ('予算', 'value'),
    ('生活スタイル', 'value'),
]

# 列名 → Excel列アルファベット
DATA_COL_LETTER = {name: get_column_letter(i + 1) for i, (name, _) in enumerate(DATA_COLUMNS)}


def formula_for(col_name: str, row_idx: int) -> str:
    """データシート派生列の数式生成（構造化参照だと数値化されない事があるため通常参照）"""
    kind = dict(DATA_COLUMNS)[col_name]
    R = row_idx

    def cell(name):
        return f'{DATA_COL_LETTER[name]}{R}'

    if kind == 'formula_gender':
        c = cell('性別コード')
        return f'=IFS({c}="1","男性",{c}="2","女性",TRUE,"未設定")'
    if kind == 'formula_reg_year':
        c = cell('登録日時')
        return f'=IF({c}="","",YEAR({c}))'
    if kind == 'formula_age':
        dob = cell('生年月日')
        reg = cell('登録日時')
        return f'=IF(OR({dob}="",{reg}=""),"",DATEDIF({dob},{reg},"Y"))'
    if kind == 'formula_age_bucket':
        age = cell('年齢')
        return (
            f'=IF({age}="","",'
            f'IF({age}<10,"",'
            f'IF({age}<=19,"10〜19",'
            f'IF({age}<=22,"20〜22",'
            f'IF({age}<=26,"23〜26",'
            f'IF({age}<=29,"27〜29",'
            f'IF({age}<=34,"30〜34",'
            f'IF({age}<=39,"35〜39",'
            f'IF({age}<=49,"40〜49",'
            f'IF({age}<=59,"50〜59","60〜"))))))))))'
        )
    if kind == 'formula_channel_flag':
        ch_name = col_name.replace('フラグ_', '')
        target = cell('きっかけ')
        return f'=IF({target}="",0,IF(ISNUMBER(SEARCH("{ch_name}",{target})),1,0))'
    return ''


def write_data_sheet(wb, records):
    ws = wb.create_sheet('データ')
    headers = [name for name, _ in DATA_COLUMNS]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = DATA_HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')

    # 値マップ
    def row_values(rec, row_idx):
        vals = []
        for name, kind in DATA_COLUMNS:
            if kind == 'value':
                if name == '店舗':
                    vals.append(rec['store'])
                elif name == '顧客ID':
                    vals.append(rec['id'])
                elif name == '氏名':
                    vals.append(rec['name'])
                elif name == '性別コード':
                    vals.append(rec['gender_code'])
                elif name == '生年月日':
                    vals.append(rec['dob'])
                elif name == '登録日時':
                    vals.append(rec['reg'])
                elif name == 'きっかけ_元データ':
                    vals.append(rec['channel_raw'])
                elif name == 'きっかけ':
                    vals.append(rec['channel_norm'])
                elif name == '決め手':
                    vals.append(rec['decision'])
                elif name == 'ペース':
                    vals.append(rec['pace'])
                elif name == '予算':
                    vals.append(rec['budget'])
                elif name == '生活スタイル':
                    vals.append(rec['lifestyle'])
                else:
                    vals.append('')
            else:
                vals.append(formula_for(name, row_idx))
        return vals

    for i, rec in enumerate(records, start=2):
        ws.append(row_values(rec, i))

    # 列幅
    widths = {
        '店舗': 8, '顧客ID': 12, '氏名': 14, '性別コード': 10, '性別': 10,
        '生年月日': 12, '登録日時': 20, '登録年度': 10, '年齢': 8, '年齢帯': 10,
        'きっかけ_元データ': 24, 'きっかけ': 24, '決め手': 24, 'ペース': 10,
        '予算': 16, '生活スタイル': 24,
    }
    for name, _ in DATA_COLUMNS:
        w = widths.get(name, 16)
        ws.column_dimensions[DATA_COL_LETTER[name]].width = w
    # 日付書式
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=len(DATA_COLUMNS)):
        for cell in row:
            col_name = DATA_COLUMNS[cell.column - 1][0]
            if col_name == '生年月日':
                cell.number_format = 'yyyy-mm-dd'
            elif col_name == '登録日時':
                cell.number_format = 'yyyy-mm-dd hh:mm:ss'

    # Excel Table として登録（追加行の自動拡張・構造化参照を有効化）
    last_col = get_column_letter(len(DATA_COLUMNS))
    last_row = len(records) + 1
    ref = f'A1:{last_col}{last_row}'
    tbl = Table(displayName=TABLE_NAME, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2', showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)

    # 先頭行固定
    ws.freeze_panes = 'A2'
    return ws


# ============ 分析シート ============
# COUNTIFS/SUMIFS の criterion 部分を組み立てる

def year_criteria(year):
    """年度フィルタ criterion。None ならフィルタなし → ('', '')。"""
    if year is None:
        return None
    return ('登録年度', year)


def build_countifs(*criteria):
    """criteria: [(列名, 値), ...] 空要素はスキップ"""
    parts = []
    for c in criteria:
        if not c:
            continue
        col, val = c
        parts.append(f'{TABLE_NAME}[{col}],"{val}"' if isinstance(val, str) else f'{TABLE_NAME}[{col}],{val}')
    return f'=COUNTIFS({",".join(parts)})'


def build_sumifs(sum_col, *criteria):
    parts = []
    for c in criteria:
        if not c:
            continue
        col, val = c
        parts.append(f'{TABLE_NAME}[{col}],"{val}"' if isinstance(val, str) else f'{TABLE_NAME}[{col}],{val}')
    return f'=SUMIFS({TABLE_NAME}[{sum_col}],{",".join(parts)})'


def build_countifs_nonblank(col, *extra):
    """指定列が非空 かつ 他の criteria も満たす件数"""
    parts = [f'{TABLE_NAME}[{col}],"<>"']
    for c in extra:
        if not c:
            continue
        col_e, val = c
        parts.append(f'{TABLE_NAME}[{col_e}],"{val}"' if isinstance(val, str) else f'{TABLE_NAME}[{col_e}],{val}')
    return f'=COUNTIFS({",".join(parts)})'


def build_gender_block(ws, start_row, year, store):
    """男女比率ブロック"""
    ws.cell(row=start_row, column=1, value='■ 男女比率')
    style_section(ws.cell(row=start_row, column=1))
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)

    for i, h in enumerate(['区分', '人数', '構成比(%)', '有効母数比(%)*'], 1):
        style_col_header(ws.cell(row=start_row + 1, column=i, value=h))

    r = start_row + 2
    yc = year_criteria(year)
    sc = ('店舗', store) if store else None

    # 男/女/未設定
    formulas = {}
    for g in GENDERS:
        formulas[g] = build_countifs(('性別', g), yc, sc)

    total_formula = build_countifs(yc, sc) if (yc or sc) else f'=ROWS({TABLE_NAME})'
    valid_formula = f'=B{r}+B{r+1}'  # 男性 + 女性

    # データ行
    labels_rows = [('男性', r), ('女性', r + 1), ('未設定', r + 2)]
    for g, row in labels_rows:
        ws.cell(row=row, column=1, value=g)
        ws.cell(row=row, column=2, value=formulas[g])
        ws.cell(row=row, column=3, value=f'=IFERROR(B{row}/B{r+3},0)')
        ws.cell(row=row, column=3).number_format = '0.0%'
        if g in ('男性', '女性'):
            ws.cell(row=row, column=4, value=f'=IFERROR(B{row}/B{r+4},0)')
            ws.cell(row=row, column=4).number_format = '0.0%'
        else:
            ws.cell(row=row, column=4, value='-')
        for c in range(1, 5):
            style_cell(ws.cell(row=row, column=c))

    # 合計行
    total_row = r + 3
    ws.cell(row=total_row, column=1, value='合計')
    ws.cell(row=total_row, column=2, value=total_formula)
    ws.cell(row=total_row, column=3, value=1.0)
    ws.cell(row=total_row, column=3).number_format = '0.0%'
    ws.cell(row=total_row, column=4, value=1.0)
    ws.cell(row=total_row, column=4).number_format = '0.0%'
    for c in range(1, 5):
        style_total(ws.cell(row=total_row, column=c))

    # 有効母数
    valid_row = r + 4
    ws.cell(row=valid_row, column=1, value='有効母数(男+女)')
    ws.cell(row=valid_row, column=2, value=valid_formula)
    ws.cell(row=valid_row, column=3, value='-')
    ws.cell(row=valid_row, column=4, value='-')
    for c in range(1, 5):
        style_cell(ws.cell(row=valid_row, column=c))

    ws.cell(row=valid_row + 1, column=1, value='* 有効母数比 = 男女合計を母数とした比率（未設定を除く）')
    ws.cell(row=valid_row + 1, column=1).font = Font(italic=True, size=9, color='7F7F7F')
    return valid_row + 3


def build_age_block(ws, start_row, year, store):
    """男女別 年齢構成"""
    ws.cell(row=start_row, column=1, value='■ 男女別 年齢構成（登録時点の満年齢）')
    style_section(ws.cell(row=start_row, column=1))
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)

    hdr = ['年齢帯', '男性', '男性 %', '女性', '女性 %', '未設定', '合計', '合計 %']
    for i, h in enumerate(hdr, 1):
        style_col_header(ws.cell(row=start_row + 1, column=i, value=h))

    yc = year_criteria(year)
    sc = ('店舗', store) if store else None

    first_data_row = start_row + 2
    for idx, (label, _, _) in enumerate(AGE_BUCKETS):
        r = first_data_row + idx
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=build_countifs(('性別', '男性'), ('年齢帯', label), yc, sc))
        ws.cell(row=r, column=4, value=build_countifs(('性別', '女性'), ('年齢帯', label), yc, sc))
        ws.cell(row=r, column=6, value=build_countifs(('性別', '未設定'), ('年齢帯', label), yc, sc))
        ws.cell(row=r, column=7, value=f'=B{r}+D{r}+F{r}')
        for c in range(1, 9):
            style_cell(ws.cell(row=r, column=c))

    # 生年月日不明行（年齢帯が空）
    r = first_data_row + len(AGE_BUCKETS)
    ws.cell(row=r, column=1, value='生年月日不明')
    ws.cell(row=r, column=2, value=build_countifs(('性別', '男性'), ('年齢帯', ''), yc, sc))
    ws.cell(row=r, column=3, value='-')
    ws.cell(row=r, column=4, value=build_countifs(('性別', '女性'), ('年齢帯', ''), yc, sc))
    ws.cell(row=r, column=5, value='-')
    ws.cell(row=r, column=6, value=build_countifs(('性別', '未設定'), ('年齢帯', ''), yc, sc))
    ws.cell(row=r, column=7, value=f'=B{r}+D{r}+F{r}')
    ws.cell(row=r, column=8, value='-')
    for c in range(1, 9):
        style_cell(ws.cell(row=r, column=c))

    # 合計行（年齢判明ベース）
    total_row = first_data_row + len(AGE_BUCKETS) + 1
    ws.cell(row=total_row, column=1, value='年齢判明合計')
    ws.cell(row=total_row, column=2, value=f'=SUM(B{first_data_row}:B{first_data_row + len(AGE_BUCKETS) - 1})')
    ws.cell(row=total_row, column=4, value=f'=SUM(D{first_data_row}:D{first_data_row + len(AGE_BUCKETS) - 1})')
    ws.cell(row=total_row, column=6, value=f'=SUM(F{first_data_row}:F{first_data_row + len(AGE_BUCKETS) - 1})')
    ws.cell(row=total_row, column=7, value=f'=B{total_row}+D{total_row}+F{total_row}')
    ws.cell(row=total_row, column=3, value=1.0)
    ws.cell(row=total_row, column=5, value=1.0)
    ws.cell(row=total_row, column=8, value=1.0)
    ws.cell(row=total_row, column=3).number_format = '0.0%'
    ws.cell(row=total_row, column=5).number_format = '0.0%'
    ws.cell(row=total_row, column=8).number_format = '0.0%'
    for c in range(1, 9):
        style_total(ws.cell(row=total_row, column=c))

    # %計算式を各行に埋め込む（合計行を分母に）
    for idx in range(len(AGE_BUCKETS)):
        r = first_data_row + idx
        ws.cell(row=r, column=3, value=f'=IFERROR(B{r}/B{total_row},0)')
        ws.cell(row=r, column=5, value=f'=IFERROR(D{r}/D{total_row},0)')
        ws.cell(row=r, column=8, value=f'=IFERROR(G{r}/G{total_row},0)')
        ws.cell(row=r, column=3).number_format = '0.0%'
        ws.cell(row=r, column=5).number_format = '0.0%'
        ws.cell(row=r, column=8).number_format = '0.0%'

    return total_row + 2


def build_channel_block(ws, start_row, year, store):
    """当店を知ったきっかけ（複数選択・フラグ列で集計）"""
    ws.cell(row=start_row, column=1, value='■ 当店を知ったきっかけ（性別別・複数選択）')
    style_section(ws.cell(row=start_row, column=1))
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)

    hdr = ['項目', '男性 件数', '男性 %', '女性 件数', '女性 %', '未設定 件数', '未設定 %', '合計 件数', '合計 %']
    for i, h in enumerate(hdr, 1):
        style_col_header(ws.cell(row=start_row + 1, column=i, value=h))

    yc = year_criteria(year)
    sc = ('店舗', store) if store else None

    # 母数行を先に計算するため位置を決める
    first = start_row + 2
    denom_row = first + len(CHANNELS)
    for idx, ch in enumerate(CHANNELS):
        r = first + idx
        flag_col = f'フラグ_{ch}'
        ws.cell(row=r, column=1, value=ch)
        ws.cell(row=r, column=2, value=build_sumifs(flag_col, ('性別', '男性'), yc, sc))
        ws.cell(row=r, column=4, value=build_sumifs(flag_col, ('性別', '女性'), yc, sc))
        ws.cell(row=r, column=6, value=build_sumifs(flag_col, ('性別', '未設定'), yc, sc))
        ws.cell(row=r, column=8, value=f'=B{r}+D{r}+F{r}')
        ws.cell(row=r, column=3, value=f'=IFERROR(B{r}/B{denom_row},0)')
        ws.cell(row=r, column=5, value=f'=IFERROR(D{r}/D{denom_row},0)')
        ws.cell(row=r, column=7, value=f'=IFERROR(F{r}/F{denom_row},0)')
        ws.cell(row=r, column=9, value=f'=IFERROR(H{r}/H{denom_row},0)')
        for c in (3, 5, 7, 9):
            ws.cell(row=r, column=c).number_format = '0.0%'
        for c in range(1, 10):
            style_cell(ws.cell(row=r, column=c))

    # 分母 = きっかけ_元データが非空 の回答者数
    ws.cell(row=denom_row, column=1, value='回答者数（%分母）')
    ws.cell(row=denom_row, column=2, value=build_countifs_nonblank('きっかけ_元データ', ('性別', '男性'), yc, sc))
    ws.cell(row=denom_row, column=4, value=build_countifs_nonblank('きっかけ_元データ', ('性別', '女性'), yc, sc))
    ws.cell(row=denom_row, column=6, value=build_countifs_nonblank('きっかけ_元データ', ('性別', '未設定'), yc, sc))
    ws.cell(row=denom_row, column=8, value=f'=B{denom_row}+D{denom_row}+F{denom_row}')
    for c in (3, 5, 7, 9):
        ws.cell(row=denom_row, column=c, value=1.0)
        ws.cell(row=denom_row, column=c).number_format = '0.0%'
    for c in range(1, 10):
        style_total(ws.cell(row=denom_row, column=c))

    r = denom_row + 1
    ws.cell(row=r, column=1, value='※ 複数選択項目。%は各性別の「回答者数」を分母（1人が複数選ぶため件数合計は100%を超える場合あり）')
    ws.cell(row=r, column=1).font = Font(italic=True, size=9, color='7F7F7F')
    return r + 2


def build_single_choice_block(ws, start_row, year, store, field_name, options, title):
    """単一選択項目（決め手/ペース/予算/生活スタイル）— 選択肢を固定リストで表示"""
    ws.cell(row=start_row, column=1, value=title)
    style_section(ws.cell(row=start_row, column=1))
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)

    hdr = ['項目', '男性 件数', '男性 %', '女性 件数', '女性 %', '未設定 件数', '未設定 %', '合計 件数', '合計 %']
    for i, h in enumerate(hdr, 1):
        style_col_header(ws.cell(row=start_row + 1, column=i, value=h))

    yc = year_criteria(year)
    sc = ('店舗', store) if store else None

    first = start_row + 2
    denom_row = first + len(options)
    for idx, opt in enumerate(options):
        r = first + idx
        ws.cell(row=r, column=1, value=opt)
        ws.cell(row=r, column=2, value=build_countifs((field_name, opt), ('性別', '男性'), yc, sc))
        ws.cell(row=r, column=4, value=build_countifs((field_name, opt), ('性別', '女性'), yc, sc))
        ws.cell(row=r, column=6, value=build_countifs((field_name, opt), ('性別', '未設定'), yc, sc))
        ws.cell(row=r, column=8, value=f'=B{r}+D{r}+F{r}')
        ws.cell(row=r, column=3, value=f'=IFERROR(B{r}/B{denom_row},0)')
        ws.cell(row=r, column=5, value=f'=IFERROR(D{r}/D{denom_row},0)')
        ws.cell(row=r, column=7, value=f'=IFERROR(F{r}/F{denom_row},0)')
        ws.cell(row=r, column=9, value=f'=IFERROR(H{r}/H{denom_row},0)')
        for c in (3, 5, 7, 9):
            ws.cell(row=r, column=c).number_format = '0.0%'
        for c in range(1, 10):
            style_cell(ws.cell(row=r, column=c))

    # 分母 = 対象列が非空 の回答数
    ws.cell(row=denom_row, column=1, value='回答合計')
    ws.cell(row=denom_row, column=2, value=build_countifs_nonblank(field_name, ('性別', '男性'), yc, sc))
    ws.cell(row=denom_row, column=4, value=build_countifs_nonblank(field_name, ('性別', '女性'), yc, sc))
    ws.cell(row=denom_row, column=6, value=build_countifs_nonblank(field_name, ('性別', '未設定'), yc, sc))
    ws.cell(row=denom_row, column=8, value=f'=B{denom_row}+D{denom_row}+F{denom_row}')
    for c in (3, 5, 7, 9):
        ws.cell(row=denom_row, column=c, value=1.0)
        ws.cell(row=denom_row, column=c).number_format = '0.0%'
    for c in range(1, 10):
        style_total(ws.cell(row=denom_row, column=c))

    r = denom_row + 1
    ws.cell(row=r, column=1, value='※ 未回答（空欄）は分母から除外。想定外の選択肢が入った場合はここに現れず、データシートで確認できます。')
    ws.cell(row=r, column=1).font = Font(italic=True, size=9, color='7F7F7F')
    return r + 2


# 選択肢マスタ（既存データから抽出）
DECISION_OPTIONS = [
    '口コミが良かった',
    'スタイル写真が好みだった',
    '予約の取りやすさ',
    'なりたいスタイルが得意そうだった',
    '駅からの近さ',
    '価格',
    'なんとなく',
    '内装・雰囲気が良さそうだった',
    '担当者を指名したかった',
]
PACE_OPTIONS = ['1ヶ月', '1.5ヶ月', '2ヶ月', '3ヶ月', '4ヶ月']
BUDGET_OPTIONS = ['〜5500円', '5500円〜7000円', '7000円〜8500円', '8500円〜10000円', '12000円以上']
LIFESTYLE_OPTIONS = [
    '土日休み', '平日休み', '休みが不定期',
    '学生', '主婦（パート・在宅で働いてる）', '主婦（家事・育児中心）', 'その他',
]


def make_analysis_sheet(wb, sheet_name, year, store, period_label):
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions['A'].width = 34
    for col in 'BCDEFGHIJK':
        ws.column_dimensions[col].width = 14

    # ヘッダ情報
    ws.cell(row=1, column=1, value=f'{STORE_NAME}  |  期間: {period_label}')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    # 対象件数
    yc = year_criteria(year)
    sc = ('店舗', store) if store else None
    if yc or sc:
        formula = build_countifs(yc, sc)
    else:
        formula = f'=ROWS({TABLE_NAME})'
    ws.cell(row=2, column=1, value='対象件数')
    ws.cell(row=2, column=2, value=formula)
    ws.cell(row=2, column=1).font = Font(bold=True)

    r = 4
    r = build_gender_block(ws, r, year, store)
    r += 1
    r = build_age_block(ws, r, year, store)
    r += 1
    r = build_channel_block(ws, r, year, store)
    r += 1
    r = build_single_choice_block(ws, r, year, store, '決め手', DECISION_OPTIONS, '■ ご来店の一番の決め手（性別別）')
    r += 1
    r = build_single_choice_block(ws, r, year, store, 'ペース', PACE_OPTIONS, '■ 今後の理想的なご来店ペース（性別別）')
    r += 1
    r = build_single_choice_block(ws, r, year, store, '予算', BUDGET_OPTIONS, '■ 本日のご予算（性別別）')
    r += 1
    r = build_single_choice_block(ws, r, year, store, '生活スタイル', LIFESTYLE_OPTIONS, '■ 生活スタイル（性別別）')
    return ws


# ============ 概要シート ============

def write_overview_sheet(wb, years):
    ws = wb.create_sheet('概要・データ状況', 0)
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 60

    ws.cell(row=1, column=1, value=f'{STORE_NAME}  顧客分析レポート')
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws.cell(row=2, column=1, value='ソース: customers_168_YYYYMMDD.csv  /  年度は登録日時ベース')
    ws.cell(row=2, column=1).font = Font(size=10, color='7F7F7F')

    ws.cell(row=4, column=1, value='■ このExcelの使い方')
    style_section(ws.cell(row=4, column=1))
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
    r = 5
    usage = [
        ('データ追加方法', '「データ」シート末尾に新規行を貼付 → Excel Tableが自動拡張 → 全分析シートが再計算'),
        ('必須列', '店舗 / 顧客ID / 性別コード(0/1/2) / 生年月日 / 登録日時 / きっかけ_元データ / きっかけ / 決め手 / ペース / 予算 / 生活スタイル'),
        ('派生列（自動計算）', '性別 / 登録年度 / 年齢 / 年齢帯 / フラグ_* — 貼付行にも数式が伝播'),
        ('「きっかけ」列の値', 'コード（例: 0_le30n6ak）ではなく日本語ラベル（ホットペッパー 等）で入力すると集計されます。CSVから追加する場合は事前に置換してください。'),
        ('店舗追加', '他店舗も「データ」に混ぜて追加可能。分析は現状 店舗=168 でフィルタ。他店舗集計を作る場合はシート複製 & 店舗IDを変更。'),
    ]
    for k, v in usage:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical='center')
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='■ データで「できること」')
    style_section(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ok_rows = [
        ('男女比率', '性別コード(0/1/2)から集計'),
        ('男女別 年齢構成', '生年月日と登録日時から算出（登録時点の満年齢）'),
        ('登録年度別 集計', '2022〜（登録日時ベース）。新年度もデータ追加で自動追加集計'),
        ('知ったきっかけ / 決め手 / 来店ペース / 予算 / 生活スタイル', 'アンケート回答者ベースで性別別に集計'),
    ]
    for k, v in ok_rows:
        c = ws.cell(row=r, column=1, value='◯')
        c.font = Font(color='2E7D32', bold=True)
        c.alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2, value=k).font = Font(bold=True)
        ws.cell(row=r, column=3, value=v)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='■ データで「できないこと」')
    style_section(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ng_rows = [
        ('3回以上／5回以上来店の分析', 'CSVに「来店回数」列がない。予約履歴／施術履歴のエクスポートが別途必要。'),
        ('全店舗合計', '現状はこのファイルに店舗168のみ。他店舗データを「データ」に追加し、必要に応じて分析シートを複製すれば対応可。'),
    ]
    for k, v in ng_rows:
        c = ws.cell(row=r, column=1, value='×')
        c.font = Font(color='C62828', bold=True)
        c.alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2, value=k).font = Font(bold=True)
        ws.cell(row=r, column=3, value=v)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='■ シート一覧')
    style_section(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    sheets_info = [('全期間', '全ての登録データ'), *[(f'{y}年度', f'{y}年 登録') for y in years]]
    for s, note in sheets_info:
        ws.cell(row=r, column=1, value=s).font = Font(bold=True)
        ws.cell(row=r, column=2, value=note)
        r += 1
    ws.cell(row=r, column=1, value='データ').font = Font(bold=True)
    ws.cell(row=r, column=2, value='生データ + 派生列。ここに行追加すれば全分析シートが更新されます。')
    return ws


def main():
    records = load_rows()
    years = sorted({r['reg'].year for r in records if r['reg']})
    print(f'Total: {len(records):,}  Years: {years}')

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    write_overview_sheet(wb, years)
    make_analysis_sheet(wb, '全期間', year=None, store=STORE_ID, period_label='全期間')
    for y in years:
        make_analysis_sheet(wb, f'{y}年度', year=y, store=STORE_ID, period_label=f'{y}年度')

    # データシートは最後に置く（分析シートで参照するのは名前ベースなので順序不問）
    write_data_sheet(wb, records)

    wb.save(OUT_PATH)
    print(f'Saved: {OUT_PATH}')


if __name__ == '__main__':
    main()
