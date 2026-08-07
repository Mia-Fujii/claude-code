"""Fill non-sensitive 基本設定 values into the enhanced workbook.

APIトークンなど機密情報は含めない。ユーザーがGoogle Sheetsで直接貼り付ける想定。
"""
import openpyxl
from openpyxl.styles import PatternFill

FILL = PatternFill("solid", fgColor="E2EFDA")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
path = "/home/user/claude-code/scratchpad/ShineALight_20ki_Enhanced.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb["基本設定"]

updates = {
    "グルコン事前フォームURL":         ("https://ws.formzu.net/dist/S16631693/", FILL),
    "ビギナーグルコン事前フォームURL": ("https://ws.formzu.net/dist/S16942447/", FILL),
    "課題作業会事前フォームURL":       ("https://ws.formzu.net/fgen/S99301863/", FILL),
    "事務局アドレス":                  ("shinealightonlineschool@gmail.com", FILL),
    "Chatwork ルームID(事務局)":       ("444552021", FILL),
    "Chatwork メンション先":           ("[To:5925962] 藤井みあ", FILL),
    "共通署名(本文末尾)":              ("---\nShine A Light 運営事務局\nMail: shinealightonlineschool@gmail.com", FILL),
    # トークンはプレースホルダのみ。実値はユーザーがGoogle Sheetsで貼り付ける。
    "Chatwork APIトークン":            ("(Google Sheetsで直接貼り付け／推奨はGASスクリプトプロパティ)", INPUT_FILL),
}
for row in ws.iter_rows(min_row=1, max_col=2):
    key = row[0].value
    if key in updates:
        val, fill = updates[key]
        row[1].value = val
        row[1].fill = fill
        print(f"[{row[0].row}] {key}")
wb.save(path)
print("saved")
