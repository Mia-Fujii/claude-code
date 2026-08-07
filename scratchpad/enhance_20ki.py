"""Enhance the user's 20期 workbook to be automation-ready.

Preserves all existing data, adds:
- 基本設定: Zoom pool, form URLs, Chatwork settings, signature
- タスク管理: ID / Zoomソース / Zoomリンク / ミーティングID / 事前フォームURL / ステータス columns
- メールテンプレート: fill all stub rows with real content (adapted from 19期), add missing event types
- 使い方: intro sheet at position 0
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from copy import copy

SRC = "/root/.claude/uploads/bc3beff8-5541-50f3-9246-cc9709efc778/2dbef318-Shine_A_Light_____.xlsx"
OUT = "/home/user/claude-code/scratchpad/ShineALight_20ki_Enhanced.xlsx"

wb = openpyxl.load_workbook(SRC)

# ---------- Styles ----------
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
INPUT_FILL  = PatternFill("solid", fgColor="FFF2CC")   # 🟡 user edits
NEW_FILL    = PatternFill("solid", fgColor="E2EFDA")   # 🟢 newly added by me
REF_FILL    = PatternFill("solid", fgColor="E7E6E6")   # 🩶 auto / reference
BODY_FONT   = Font(name="Arial", size=10)
BOLD        = Font(name="Arial", bold=True, size=10)
WRAP        = Alignment(wrap_text=True, vertical="top")
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# =============================================================
# 1) 使い方 sheet (new, at position 0)
# =============================================================
ws0 = wb.create_sheet("使い方", 0)
ws0.column_dimensions["A"].width = 3
ws0.column_dimensions["B"].width = 26
ws0.column_dimensions["C"].width = 100

ws0["B2"] = "Shine A Light 20期 自動化マスタ"
ws0["B2"].font = Font(name="Arial", bold=True, size=16, color="1F3864")

intro = [
    ("", ""),
    ("このファイルの位置づけ",
     "20期の講座運営を、Gmail下書き自動作成＋Chatwork通知＋フォーム自動開閉で\n"
     "回すためのマスタデータです。GAS（Google Apps Script）がこのファイルを毎朝読んで動きます。"),
    ("", ""),
    ("シート構成", ""),
    ("① 基本設定",
     "全メール共通で使う値（会員サイト、Zoomプール、フォームURL、Chatwork設定、署名 など）。\n"
     "期をまたいで基本変わらない値の置き場。"),
    ("② タスク管理",
     "20期の全イベントを1行1件で管理（29件入力済み）。\n"
     "GASが毎朝ここを読んで「今日は各イベントの何のメール送信日か」を判定します。"),
    ("③ メールテンプレート",
     "イベント種別 × タイミング（3日前／前日／当日／アーカイブ など）ごとの雛形。\n"
     "{{差込項目}} が自動置換されます。"),
    ("", ""),
    ("色分けルール", ""),
    ("🟡 黄色セル", "あなたが埋める場所（Zoomリンク、フォームURL、Chatwork設定など）"),
    ("🟢 緑セル", "今回私が追加した列／行（既存データを触らず追加）"),
    ("🩶 灰色セル", "自動計算・参照専用（触らない）"),
    ("⬜ 白セル", "既存のデータ（あなたが元々埋めていた場所）"),
    ("", ""),
    ("最初にやること", ""),
    ("① 基本設定 の🟡を埋める",
     "・グルコン共通ZoomURL＋ミーティングID（若菜先生のグルコン専用リンク）\n"
     "・キャッチアップマンデー共通ZoomURL＋ID（若菜先生のマンデー専用リンク）\n"
     "・グルコン／ビギナー／課題作業会 事前フォームURL（3本）\n"
     "・Chatwork APIトークン＋ルームID＋メンション先\n"
     "・送信元アドレス"),
    ("② タスク管理 のZoomソース列（🟢）を確認",
     "自動で設定済み。以下ルールで自動判定しています：\n"
     "・グルコン → 「若菜グルコン共通」\n"
     "・キャッチアップマンデー → 「若菜マンデー共通」\n"
     "・キャッチアップウェンズデー／サポート講師ビギナーグルコン → 「未定（サポート講師待ち）」\n"
     "・オリエンテーション／課題作業会 → 「個別」（別途Zoomリンク列に直接入力）\n"
     "・動画配信 → 「Zoomなし」"),
    ("③ メールテンプレート を読んで直したい部分だけ調整",
     "19期の実データを元に本文込みで埋めておきました。\n"
     "文言を変えたい箇所があれば直してOK（既存のオリエンテーション3本はそのまま残しています）。"),
    ("", ""),
    ("Step 2（次回）",
     "このファイルがGoogleスプレッドシートにアップロードされ、内容が確認できたら、\n"
     "GASコードを書きます。まずはグルコン系（一番数が多い）だけで動かして、動作確認後に他を追加。"),
]
r = 3
for lbl, val in intro:
    ws0.cell(row=r, column=2, value=lbl).font = Font(name="Arial", bold=True, size=11, color="1F3864")
    c = ws0.cell(row=r, column=3, value=val)
    c.font = BODY_FONT; c.alignment = WRAP
    ws0.row_dimensions[r].height = 70 if val and val.count("\n") >= 2 else (45 if "\n" in val else 22)
    r += 1

# =============================================================
# 2) 基本設定 sheet enhancements
# =============================================================
ws1 = wb["基本設定"]

# find the last non-empty row in cols A-B
last = 1
for row in ws1.iter_rows(min_row=1, max_col=2, values_only=False):
    if row[0].value or row[1].value:
        last = row[0].row

# make sure column widths are readable
ws1.column_dimensions["A"].width = 32
ws1.column_dimensions["B"].width = 70
ws1.column_dimensions["C"].width = 30

# add a section header + new keys after existing content
r = last + 2
def add_setting(key, value, note="", fill=NEW_FILL):
    global r
    ws1.cell(row=r, column=1, value=key).font = BOLD
    ws1.cell(row=r, column=1).alignment = WRAP; ws1.cell(row=r, column=1).border = BORDER
    c = ws1.cell(row=r, column=2, value=value)
    c.font = BODY_FONT; c.alignment = WRAP; c.border = BORDER; c.fill = fill
    if note:
        n = ws1.cell(row=r, column=3, value=note)
        n.font = Font(name="Arial", italic=True, color="595959", size=9)
        n.alignment = WRAP; n.border = BORDER
    r += 1

def add_section_title(title):
    global r
    ws1.cell(row=r, column=1, value=title).font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws1.cell(row=r, column=1).fill = HEADER_FILL
    ws1.cell(row=r, column=1).alignment = CENTER
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

add_section_title("── Zoomリンク・プール（共通Zoom用）──")
add_setting("グルコン共通ZoomURL",             "https://us02web.zoom.us/j/83540214061", "若菜先生グルコン全回で使う1本")
add_setting("グルコン共通ミーティングID",       "835 4021 4061", "")
add_setting("キャッチアップマンデー共通ZoomURL","https://us02web.zoom.us/j/84912148179", "若菜先生マンデー全回で使う1本")
add_setting("キャッチアップマンデー共通ミーティングID","849 1214 8179", "")
add_setting("オリエンテーションZoomURL",        "https://us02web.zoom.us/j/82603439180", "既存テンプレから移設")
add_setting("オリエンテーションミーティングID", "826 0343 9180", "")

add_section_title("── 事前フォームURL ──")
add_setting("グルコン事前フォームURL",         "", "🟡 若菜グルコン用アンケート")
add_setting("ビギナーグルコン事前フォームURL", "", "🟡 サポート講師ビギナーグルコン用")
add_setting("課題作業会事前フォームURL",       "", "🟡")

add_section_title("── Chatwork設定 ──")
add_setting("Chatwork APIトークン",   "", "🟡 Chatwork設定＞API から取得")
add_setting("Chatwork ルームID(事務局)","", "🟡 通知投稿先ルームのID")
add_setting("Chatwork メンション先",  "", "🟡 例：[To:1234567] みあさん")

add_section_title("── 差出人・署名 ──")
add_setting("事務局差出人名", "Shine A Light 運営事務局", "")
add_setting("事務局アドレス", "", "🟡 実際の送信元アドレス")
add_setting("共通署名(本文末尾)", "---\nShine A Light 運営事務局\nMail: your-office@example.com", "🟡 実運用に合わせて調整")

add_section_title("── GAS実行時刻 ──")
add_setting("下書き作成時刻", "07:00", "毎朝この時刻にGASが動く")

# =============================================================
# 3) タスク管理 sheet enhancements
# =============================================================
ws2 = wb["タスク管理"]

# existing headers are in row 1, columns A..K (11 columns)
# find the actual data range (non-empty rows in col A)
last_row = 1
for row in ws2.iter_rows(min_row=2, max_col=1, values_only=False):
    if row[0].value is not None:
        last_row = row[0].row

# find last used column (existing sheet uses A..K)
last_col = 11  # 内容 through 会員サイト（アーカイブ用）

# Add new columns starting at column L
new_headers = [
    "ID", "Zoomソース", "Zoomリンク", "ミーティングID", "事前フォームURL", "ステータス"
]
for i, h in enumerate(new_headers):
    col = last_col + 1 + i
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER
    ws2.column_dimensions[get_column_letter(col)].width = [10, 22, 40, 18, 40, 12][i]

# Read event types from column A and assign Zoomソース auto
def guess_zoom_source(content):
    if content is None: return ""
    c = str(content)
    if "オリエンテーション" in c: return "個別"
    if "動画配信" in c: return "Zoomなし"
    if "キャッチアップマンデー" in c: return "若菜マンデー共通"
    if "キャッチアップウェンズデー" in c: return "未定(サポート講師待ち)"
    if "サポート講師ビギナーグルコン" in c: return "未定(サポート講師待ち)"
    if "課題作業会" in c: return "個別"
    if c.strip() == "グルコン" or c.strip().endswith("グルコン"):
        # exclude ビギナーグルコン (caught above)
        return "若菜グルコン共通"
    return ""

def guess_form_url(content):
    if content is None: return ""
    c = str(content)
    if "オリエンテーション" in c: return ""
    if "動画配信" in c: return ""
    if "キャッチアップ" in c: return ""
    if "サポート講師ビギナーグルコン" in c: return "={{ビギナーグルコン事前フォームURL}}"
    if "課題作業会" in c: return "={{課題作業会事前フォームURL}}"
    if c.strip() == "グルコン" or c.strip().endswith("グルコン"):
        return "={{グルコン事前フォームURL}}"
    return ""

# fill new columns for each existing row
event_seq = 0
for r_idx in range(2, last_row + 1):
    content = ws2.cell(row=r_idx, column=1).value
    if content is None: continue
    event_seq += 1
    # ID
    id_cell = ws2.cell(row=r_idx, column=last_col+1, value=f"E{event_seq:03d}")
    id_cell.font = BODY_FONT; id_cell.fill = REF_FILL; id_cell.border = BORDER
    id_cell.alignment = WRAP
    # Zoomソース
    zs = guess_zoom_source(content)
    zs_cell = ws2.cell(row=r_idx, column=last_col+2, value=zs)
    zs_cell.font = BODY_FONT; zs_cell.fill = NEW_FILL; zs_cell.border = BORDER
    zs_cell.alignment = WRAP
    # Zoomリンク (blank; user fills for 個別/未定)
    zl_cell = ws2.cell(row=r_idx, column=last_col+3)
    zl_cell.font = BODY_FONT; zl_cell.fill = INPUT_FILL; zl_cell.border = BORDER
    zl_cell.alignment = WRAP
    # For オリエンテーション row, pre-fill from settings
    if content and "オリエンテーション" in str(content):
        zl_cell.value = "https://us02web.zoom.us/j/82603439180"
        ws2.cell(row=r_idx, column=last_col+4, value="826 0343 9180")
    # ミーティングID
    mid = ws2.cell(row=r_idx, column=last_col+4)
    mid.font = BODY_FONT; mid.fill = INPUT_FILL; mid.border = BORDER; mid.alignment = WRAP
    # フォームURL (as marker text; GAS resolves)
    furl_val = guess_form_url(content)
    furl = ws2.cell(row=r_idx, column=last_col+5, value=furl_val)
    furl.font = BODY_FONT; furl.fill = NEW_FILL; furl.border = BORDER; furl.alignment = WRAP
    # ステータス
    st = ws2.cell(row=r_idx, column=last_col+6, value="未実施")
    st.font = BODY_FONT; st.fill = INPUT_FILL; st.border = BORDER; st.alignment = WRAP

# freeze first row
ws2.freeze_panes = "A2"

# add comment on Zoomソース header
ws2.cell(row=1, column=last_col+2).comment = Comment(
    "選択肢：若菜グルコン共通 / 若菜マンデー共通 / 個別 / 未定(サポート講師待ち) / Zoomなし\n"
    "GASはこの値を見てZoomリンクの取得元を決めます。",
    "system"
)
ws2.cell(row=1, column=last_col+5).comment = Comment(
    "={{変数名}} は基本設定シートの該当値を自動参照。\n"
    "直接URLを書くこともできます。",
    "system"
)

# =============================================================
# 4) メールテンプレート sheet - fill stub rows and add missing
# =============================================================
ws3 = wb["メールテンプレート"]

# Preserve existing templates that are complete (rows 2-4: オリエンテーション)
# Overwrite stubs (rows 5+)
# ensure headers: 講義内容 | タイミング | 件名 | 本文 | zoomリンク
# We'll extend the model: also fill zoomリンク with a resolver marker

# Real templates adapted from 19期 data, using 20期 variable names ({{日程}} {{期}} {{zoomリンク}} etc.)
CLOSING = ("ご案内は以上です。\n\n"
           "ご不明点のお問い合わせは\n"
           "当メールの返信にてご連絡ください。\n\n"
           "Shine A Light 運営事務局")

GRULCON_ADDENDUM = """【グルコンで添削をご希望の場合の注意点】

■ 対象物：スライド／SNS（プロフィール含む）／LP／ホームページ／Googleドキュメント など
■ 提出方法：事前質問フォームに、対象物のURLリンクを必ず貼り付け。閲覧権限は「リンクを知っている全員が閲覧可」に設定。
■ 質問の書き方：「見てほしいです」の一言でなく、特に見てほしい箇所／現状と目的／相談内容 を具体的にお書きください。
■ 受付できないケース：リンク未記載／個別メッセージでの直前送付／閲覧権限が付与されていないリンク"""

templates_add = [
    # (講義内容, タイミング, 件名, 本文)
    # -- 動画配信 (6回分の当日) --
    ("第1回動画配信", "当日",
     "【重要】第1回目の講義動画をお届けいたします",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは\nShine A Light 運営事務局です。\n\n"
     "本日、第1回目の講義動画を会員サイトにアップいたしました。\n\n"
     "＝＝＝＝＝\n\n{{会員サイト}}\n\n＝＝＝＝＝\n\n"
     "モジュールの項目で第1回目の講義動画をご視聴可能です。\n\n"
     "＜第1回目：ブランディング＞\n"
     "①３ヶ月でオンラインで成功するためのマインドセット\n"
     "②ストーリーの作り方\n"
     "③実績を掘り起こそう／お客様のお声を取ってみよう\n"
     "④魅力的なプロフィール作成\n"
     "⑤Youtubeで市場調査\n\n"
     "課題提出の期限は次回の講義動画配信日までとなっております。\n\n"
     + CLOSING),
    ("第2回動画配信", "当日",
     "【重要】第2回目の講義動画をお届けいたします",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは\nShine A Light 運営事務局です。\n\n"
     "本日、第2回目の講義動画を会員サイトにアップいたしました。\n\n"
     "＝＝＝＝＝\n\n{{会員サイト}}\n\n＝＝＝＝＝\n\n"
     "モジュールの項目でご視聴ください。\n\n"
     "課題提出の期限は次回の講義動画配信日までとなっております。\n\n"
     + CLOSING),
    ("第3回動画配信", "当日",
     "【重要】第3回目の講義動画をお届けいたします",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは\nShine A Light 運営事務局です。\n\n"
     "本日、第3回目の講義動画を会員サイトにアップいたしました。\n\n"
     "＝＝＝＝＝\n\n{{会員サイト}}\n\n＝＝＝＝＝\n\n"
     "課題提出の期限は次回の講義動画配信日までとなっております。\n\n"
     + CLOSING),
    ("第4回動画配信", "当日",
     "【重要】第4回目の講義動画をお届けいたします",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは\nShine A Light 運営事務局です。\n\n"
     "本日、第4回目の講義動画を会員サイトにアップいたしました。\n\n"
     "＝＝＝＝＝\n\n{{会員サイト}}\n\n＝＝＝＝＝\n\n"
     "課題提出の期限は次回の講義動画配信日までとなっております。\n\n"
     + CLOSING),
    ("第5回動画配信", "当日",
     "【重要】第5回目の講義動画をお届けいたします",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは\nShine A Light 運営事務局です。\n\n"
     "本日、第5回目の講義動画を会員サイトにアップいたしました。\n\n"
     "＝＝＝＝＝\n\n{{会員サイト}}\n\n＝＝＝＝＝\n\n"
     "課題提出の期限は次回の講義動画配信日までとなっております。\n\n"
     + CLOSING),
    ("第6回動画配信", "当日",
     "【重要】第6回目の講義動画をお届けいたします",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは\nShine A Light 運営事務局です。\n\n"
     "本日、第6回目の講義動画を会員サイトにアップいたしました。\n\n"
     "＝＝＝＝＝\n\n{{会員サイト}}\n\n＝＝＝＝＝\n\n"
     "3ヶ月お疲れ様でした！\n\n"
     + CLOSING),

    # -- キャッチアップウェンズデー --
    ("キャッチアップウェンズデー", "メールセット",
     "【重要】{{日程}}キャッチアップウェンズデーのご案内",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\nShine A Light 運営事務局です。\n\n"
     "{{日程}}{{時間}}にサポート講師による\nキャッチアップウェンズデーを開催します。\n\n"
     "今回のサポート講師は{{担当者}}です。\n\n"
     "気軽にご参加ください。\n\n"
     + CLOSING),
    ("キャッチアップウェンズデー", "前日",
     "【重要】明日{{時間}}キャッチアップウェンズデー詳細",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\nShine A Light 運営事務局です。\n\n"
     "明日 {{日程}}{{時間}} にキャッチアップウェンズデーがございます。\n\n"
     "＝＝＝＝＝＝\n\n【参加URL（Zoom）】\n{{zoomリンク}}\n\n＝＝＝＝＝＝\n\n"
     "サポート講師：{{担当者}}\n\n"
     + CLOSING),
    ("キャッチアップウェンズデー", "当日",
     "【重要】本日{{時間}}キャッチアップウェンズデー詳細",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\nShine A Light 運営事務局です。\n\n"
     "本日 {{日程}}{{時間}} にキャッチアップウェンズデーがございます。\n\n"
     "＝＝＝＝＝＝\n\n【参加URL（Zoom）】\n{{zoomリンク}}\n\n＝＝＝＝＝＝\n\n"
     "サポート講師：{{担当者}}\n\n"
     "それでは後ほど、よろしくお願いいたします。\n\n"
     + CLOSING),
    ("キャッチアップウェンズデー", "アーカイブ",
     "【アーカイブ】{{日程}}キャッチアップウェンズデーの動画です",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\nShine A Light 運営事務局です。\n\n"
     "昨日開催しましたキャッチアップウェンズデーの動画を\n会員サイトにアップしました。\n\n"
     "参加できなかった方、復習等にご活用ください。\n\n"
     "＝＝＝＝＝＝\n\n{{会員サイト}}\n\n＝＝＝＝＝＝\n\n"
     + CLOSING),

    # -- キャッチアップマンデー --
    ("キャッチアップマンデー", "メールセット",
     "【重要】{{日程}}キャッチアップマンデーのご案内",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\nShine A Light 運営事務局です。\n\n"
     "{{日程}}{{時間}}に\n若菜先生のキャッチアップマンデーを開催します。\n\n"
     "気軽にご参加ください。\n\n"
     + CLOSING),
    ("キャッチアップマンデー", "前日",
     "【重要】明日{{時間}}キャッチアップマンデー詳細",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\nShine A Light 運営事務局です。\n\n"
     "明日 {{日程}}{{時間}} にキャッチアップマンデーがございます。\n\n"
     "＝＝＝＝＝＝\n\n【参加URL（Zoom）】\n{{zoomリンク}}\n\n＝＝＝＝＝＝\n\n"
     "担当：{{担当者}}\n\n"
     + CLOSING),
    ("キャッチアップマンデー", "当日",
     "【重要】本日{{時間}}キャッチアップマンデー詳細",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\nShine A Light 運営事務局です。\n\n"
     "本日 {{日程}}{{時間}} にキャッチアップマンデーがございます。\n\n"
     "＝＝＝＝＝＝\n\n【参加URL（Zoom）】\n{{zoomリンク}}\n\n＝＝＝＝＝＝\n\n"
     "担当：{{担当者}}\n\n"
     "それでは後ほど、よろしくお願いいたします。\n\n"
     + CLOSING),
    ("キャッチアップマンデー", "アーカイブ",
     "【アーカイブ】{{日程}}キャッチアップマンデーの動画です",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\nShine A Light 運営事務局です。\n\n"
     "昨日のキャッチアップマンデーのアーカイブを\n会員サイトにアップしました。\n\n"
     "参加できなかった方、復習等にご活用ください。\n\n"
     "＝＝＝＝＝＝\n\n{{会員サイト}}\n\n＝＝＝＝＝＝\n\n"
     + CLOSING),

    # -- サポート講師ビギナーグルコン --
    ("サポート講師ビギナーグルコン", "メールセット",
     "【重要】{{日程}}{{時間}}開催 ビギナーグルコンの質問をお送りください",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\n\"Shine A Light\"運営事務局です。\n\n"
     "{{日程}}{{時間}}にサポート講師によるビギナーグルコンを開催します。\n\n"
     "今回のサポート講師は{{担当者}}です。\n\n"
     "困っていること・相談したいことがありましたら、何でも小さなことでもいいのでお送りください。\n\n"
     "【グルコン開催前の事前アンケートのお願い】\n\n"
     "以下のアンケートフォームから送信してください。\n"
     "いただいた質問にグルコンでお答えいたします。\n\n"
     "▼グルコンの事前アンケート▼\n{{事前フォームURL}}\n\n"
     + CLOSING),
    ("サポート講師ビギナーグルコン", "3日前",
     "【リマインド】{{日程}}ビギナーグルコンの事前質問はお済みですか？",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\n\"Shine A Light\"運営事務局です。\n\n"
     "{{日程}}{{時間}}のビギナーグルコンまであと3日です。\n\n"
     "▼事前アンケートフォーム▼\n{{事前フォームURL}}\n\n"
     "サポート講師：{{担当者}}\n\n"
     + CLOSING),
    ("サポート講師ビギナーグルコン", "前日",
     "【重要】明日{{時間}}〜ビギナーグルコン詳細・事前質問受付は本日12時まで！",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\n\"Shine A Light\"運営事務局です。\n\n"
     "明日のビギナーグルコンの詳細と、事前質問アンケート送信のお願いです。\n\n"
     "明日 {{日程}}{{時間}} にビギナーグルコンがございます。\n\n"
     "＝＝＝＝＝＝\n\n【参加URL（Zoom）】\n{{zoomリンク}}\n\n＝＝＝＝＝＝\n\n"
     "【事前質問受付は本日12時まで】\n{{事前フォームURL}}\n\n"
     "サポート講師：{{担当者}}\n\n"
     + CLOSING),
    ("サポート講師ビギナーグルコン", "当日",
     "【重要】本日{{時間}}〜ビギナーグルコン詳細",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\n\"Shine A Light\"運営事務局です。\n\n"
     "本日 {{日程}}{{時間}} にビギナーグルコンがございます。\n\n"
     "＝＝＝＝＝＝\n\n【参加URL（Zoom）】\n{{zoomリンク}}\n\n＝＝＝＝＝＝\n\n"
     "サポート講師：{{担当者}}\n\n"
     "それでは後ほど、よろしくお願いいたします。\n\n"
     + CLOSING),
    ("サポート講師ビギナーグルコン", "アーカイブ",
     "【アーカイブ】{{日程}}ビギナーグルコンの動画です",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\n\"Shine A Light\"運営事務局です。\n\n"
     "昨日のビギナーグルコンのアーカイブを会員サイトにアップしました。\n\n"
     "参加できなかった方や復習にご活用ください。\n\n"
     "＝＝＝＝＝＝\n\n{{会員サイト}}\n\n＝＝＝＝＝＝\n\n"
     + CLOSING),

    # -- グルコン (若菜) --
    ("グルコン", "メールセット",
     "{{日程}}{{時間}}〜開催グルコンの質問をお送りください",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\n\"Shine A Light\"運営事務局です。\n\n"
     "{{日程}}{{時間}}〜開催のグルコンについてのご連絡です。\n\n"
     "課題をこなすのに、前へ進むのに、止まってしまっていませんか？\n\n"
     "グルコンへの質問を受け付けますので、\n"
     "ぜひ困っていること・相談したいことがありましたら、\n"
     "何でも小さなことでもいいのでお送りください。\n\n"
     "【グルコン開催前の事前アンケートのお願い】\n\n"
     "▼グルコンの事前アンケート▼\n{{事前フォームURL}}\n\n"
     "※グルコンの欠席連絡は不要です。\n\n" + GRULCON_ADDENDUM + "\n\n" + CLOSING),
    ("グルコン", "3日前",
     "【リマインド】{{日程}}グルコン 事前質問はお済みですか？",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\n\"Shine A Light\"運営事務局です。\n\n"
     "{{日程}}{{時間}}のグルコンまであと3日です。\n\n"
     "事前質問がまだの方はぜひお送りください。\n\n"
     "▼グルコンの事前アンケート▼\n{{事前フォームURL}}\n\n"
     + CLOSING),
    ("グルコン", "前日",
     "【重要】明日{{時間}}〜グルコン詳細・事前質問受付は本日12時まで！",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\n\"Shine A Light\"運営事務局です。\n\n"
     "明日のグルコンの詳細のお知らせと事前質問アンケート送信のお願いです。\n\n"
     "明日 {{日程}}{{時間}} にグルコンがございます。\n\n"
     "＝＝＝＝＝＝\n\n【グルコン参加URL（Zoom）】\n{{zoomリンク}}\n\n＝＝＝＝＝＝\n\n"
     "※グルコン動画は、後日アーカイブを送付いたします。\n"
     "※グルコンの欠席連絡は不要です。\n\n"
     "【グルコン事前質問受付は本日12時まで！】\n{{事前フォームURL}}\n\n"
     + GRULCON_ADDENDUM + "\n\n" + CLOSING),
    ("グルコン", "当日",
     "【重要】本日{{時間}}〜グルコンの詳細です！",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\n\"Shine A Light\"運営事務局です。\n\n"
     "本日 {{日程}}{{時間}} にグルコンがございます。\n\n"
     "＝＝＝＝＝＝\n\n【グルコン参加URL（Zoom）】\n{{zoomリンク}}\n\n＝＝＝＝＝＝\n\n"
     "※グルコン動画は、後日アーカイブを送付いたします。\n"
     "本日ご都合がつかない方は、アーカイブをご活用ください。\n\n"
     "それでは後ほど、よろしくお願いいたします。\n\n" + CLOSING),
    ("グルコン", "アーカイブ",
     "【グルコン】{{日程}}開催のアーカイブ動画です！",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\n\"Shine A Light\"運営事務局です。\n\n"
     "昨日のグルコンのアーカイブを会員サイトにアップしました！\n\n"
     "参加できなかった方や復習にご活用くださいませ。\n\n"
     "＝＝＝＝＝＝\n\n{{会員サイト}}\n\n＝＝＝＝＝＝\n\n"
     + CLOSING),

    # -- 課題作業会 --
    ("課題作業会", "メールセット",
     "【重要】{{日程}}{{時間}}開催 課題作業会の質問をお送りください",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\n\"Shine A Light\"運営事務局です。\n\n"
     "{{日程}}{{時間}}に課題作業会を開催します。\n\n"
     "「作業会でこんな作業を一緒にしてほしい！」\n"
     "「〇〇に困っている！」\n"
     "ということがありましたら、事前にお伺いしますので\n事前フォームに入力ください。\n\n"
     "▼作業会の事前フォーム▼\n{{事前フォームURL}}\n\n"
     "《作業会の詳細》\n【日時】{{日程}}{{時間}}\n・自由参加 ・途中参加、途中退出OK\n【担当講師】{{担当者}}\n\n"
     + CLOSING),
    ("課題作業会", "3日前",
     "【リマインド】{{日程}}課題作業会 事前質問はお済みですか？",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\n\"Shine A Light\"運営事務局です。\n\n"
     "{{日程}}{{時間}}の課題作業会まであと3日です。\n\n"
     "▼事前フォーム▼\n{{事前フォームURL}}\n\n"
     + CLOSING),
    ("課題作業会", "前日",
     "【重要】明日{{時間}}〜作業会詳細・事前質問受付は本日12時まで！",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\n\"Shine A Light\"運営事務局です。\n\n"
     "明日の作業会の詳細と事前フォーム送信のお願いです。\n\n"
     "明日 {{日程}}{{時間}} に作業会がございます。\n\n"
     "＝＝＝＝＝＝\n\n【参加URL（Zoom）】\n{{zoomリンク}}\n\n＝＝＝＝＝＝\n\n"
     "【担当講師】{{担当者}}\n\n"
     "【事前質問の回答期限は本日12時まで！】\n{{事前フォームURL}}\n\n"
     + CLOSING),
    ("課題作業会", "当日",
     "【重要】本日{{時間}}〜作業会の詳細です！",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\n\"Shine A Light\"運営事務局です。\n\n"
     "本日 {{日程}}{{時間}} に作業会がございます。\n\n"
     "＝＝＝＝＝＝\n\n【参加URL（Zoom）】\n{{zoomリンク}}\n\n＝＝＝＝＝＝\n\n"
     "《作業会の詳細》\n【担当講師】{{担当者}}\n\n"
     "それでは後ほど、よろしくお願いいたします。\n\n" + CLOSING),
    ("課題作業会", "アーカイブ",
     "【作業会】{{日程}}開催のアーカイブ動画です！",
     "※このメールは、『{{期}}\"Shine A Light\"』の皆さまにお送りしております。\n\n"
     "こんにちは。\n\"Shine A Light\"運営事務局です。\n\n"
     "昨日の作業会のアーカイブを会員サイトにアップしました！\n\n"
     "参加できなかった方や復習にご活用くださいませ。\n\n"
     "＝＝＝＝＝＝\n\n{{会員サイト}}\n\n＝＝＝＝＝＝\n\n"
     + CLOSING),
]

# Find where to start writing new template rows: replace stubs from row 5 onward, then append
# First, wipe existing rows 5..17 (stub rows) so we can rewrite cleanly
for r_idx in range(5, 18):
    for c_idx in range(1, 6):
        ws3.cell(row=r_idx, column=c_idx).value = None

# Set column widths
ws3.column_dimensions["A"].width = 26
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 55
ws3.column_dimensions["D"].width = 90
ws3.column_dimensions["E"].width = 30

# Freeze
ws3.freeze_panes = "A2"

# Restyle headers
for c in range(1, 6):
    cell = ws3.cell(row=1, column=c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

# Write templates starting at row 5
r_idx = 5
for content, timing, subject, body in templates_add:
    ws3.cell(row=r_idx, column=1, value=content)
    ws3.cell(row=r_idx, column=2, value=timing)
    ws3.cell(row=r_idx, column=3, value=subject)
    ws3.cell(row=r_idx, column=4, value=body)
    # zoomリンク column stays empty (resolved by GAS from タスク管理.Zoomソース)
    ws3.cell(row=r_idx, column=5, value="")
    for c in range(1, 6):
        cell = ws3.cell(row=r_idx, column=c)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        cell.fill = NEW_FILL
    ws3.row_dimensions[r_idx].height = 180
    r_idx += 1

# Add a comment on the zoomリンク column header
ws3.cell(row=1, column=5).comment = Comment(
    "テンプレ内の {{zoomリンク}} は、GAS実行時に「タスク管理」シートのZoomソースに従って\n"
    "自動で解決されます（若菜グルコン共通 / 若菜マンデー共通 / 個別 / 未定）。\n"
    "この列は基本空でOK。テンプレ固有のURLを埋め込みたいときだけ使う。",
    "system"
)

wb.save(OUT)
print("wrote", OUT)
