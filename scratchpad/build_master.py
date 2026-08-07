"""Build the Shine A Light master template workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

OUT = "/home/user/claude-code/scratchpad/ShineALight_Master.xlsx"

# ---------- Styles ----------
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
INPUT_FILL  = PatternFill("solid", fgColor="FFF2CC")   # yellow = user edits
REF_FILL    = PatternFill("solid", fgColor="E7E6E6")   # gray  = do not edit / auto
LEGEND_FILL = PatternFill("solid", fgColor="D9E1F2")
NOTE_FONT   = Font(name="Arial", italic=True, color="595959", size=10)
BODY_FONT   = Font(name="Arial", size=10)
WRAP        = Alignment(wrap_text=True, vertical="top")
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def fill_row(ws, row, values, fill=None):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        if fill:
            cell.fill = fill

wb = openpyxl.Workbook()

# =================================================================
# Sheet 0: 使い方
# =================================================================
ws0 = wb.active
ws0.title = "使い方"
set_widths(ws0, [4, 24, 90])

ws0["B2"] = "Shine A Light 事務局オートメーション マスタ"
ws0["B2"].font = Font(name="Arial", bold=True, size=16, color="1F3864")
ws0.merge_cells("B2:C2")

intro = [
    ("", ""),
    ("目的", "GAS（Google Apps Script）が毎朝このスプレッドシートを読み、\n各イベントの何日前かを判定 → Gmail下書きを自動生成 → Chatworkに通知します。"),
    ("編集する場所", "🟡 黄色セル … あなたが埋める場所（イベント日程・Zoomリンク・フォームURLなど）\n⬜ 白セル …… システムが読む固定値（テンプレ本文、共通URLなど・必要なら編集可）\n🩶 灰色セル … 自動計算 or 参照専用（触らない）"),
    ("", ""),
    ("シート構成", ""),
    ("① イベント日程", "全イベント（グルコン／作業会／講義配信／締切系など）を1行1件で管理。\n新しい期が始まったら、このシートに追記するだけでOK。"),
    ("② 共通設定", "会員サイトURL・FBグループURL・署名・事務局アドレス・Chatwork設定など、\n全メール共通で使う値。期を跨いで基本変わらない。"),
    ("③ 本文テンプレ", "「質問募集」「前日」「当日」「アーカイブ」など、メール種別ごとの件名＆本文の雛形。\n{{差込項目}} が自動置換される。"),
    ("④ 差込項目 リファレンス", "テンプレで使える {{変数名}} の一覧。テンプレ編集時に参照。"),
    ("", ""),
    ("運用の流れ", "1. 期の頭に「① イベント日程」を1回埋める（20期なら20期分を一気に）\n2. GASが毎朝走って、「今日は◯◯の3日前か？」を判定\n3. 該当あり → テンプレ選ぶ → 差込 → Gmail下書き作成 → Chatwork通知\n4. あなた：Chatwork通知が来たら Gmail下書きを確認して送信ボタン"),
    ("", ""),
    ("Step 2以降", "このスプレッドシートが埋まったら、GASコードを書きます（Step 2）。\nまずはこのファイルを Google スプレッドシートにアップロードして、\n黄色セルを実データで埋めてみてください。"),
]
r = 3
for label, val in intro:
    ws0.cell(row=r, column=2, value=label).font = Font(name="Arial", bold=True, size=11, color="1F3864")
    c = ws0.cell(row=r, column=3, value=val)
    c.font = BODY_FONT
    c.alignment = WRAP
    ws0.row_dimensions[r].height = 60 if val and "\n" in val else 20
    r += 1

# =================================================================
# Sheet 1: イベント日程
# =================================================================
ws1 = wb.create_sheet("① イベント日程")

headers1 = [
    "イベントID", "期", "イベント種別", "イベント名（表示用）",
    "開催日", "曜日(自動)", "開始時刻", "終了時刻",
    "担当講師", "対象セグメント",
    "Zoomリンク", "ミーティングID",
    "事前フォームURL", "フォーム開放日", "フォーム締切日時",
    "アーカイブURL", "備考", "ステータス"
]
widths1 = [10, 6, 18, 28, 12, 8, 10, 10, 18, 24, 40, 16, 40, 12, 18, 40, 30, 12]
set_widths(ws1, widths1)
for i, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=i, value=h)
style_header(ws1, 1, len(headers1))
ws1.freeze_panes = "A2"
ws1.row_dimensions[1].height = 32

# comments on some headers
ws1.cell(row=1, column=1).comment = Comment("システムが自動発番。手入力しない。", "system")
ws1.cell(row=1, column=3).comment = Comment("グルコン / ビギナーグルコン / 課題作業会 / 講義配信 / キャッチアップサンデー / 締切系 / 開講準備 のいずれか", "system")
ws1.cell(row=1, column=6).comment = Comment("開催日から自動計算（=TEXT(E2,\"aaa\")）", "system")
ws1.cell(row=1, column=10).comment = Comment("19期 / 19期継続 / 19期継続(作業会特典対象者) など", "system")
ws1.cell(row=1, column=14).comment = Comment("この日付の朝にフォームを自動オープン。空欄なら開閉しない。", "system")
ws1.cell(row=1, column=15).comment = Comment("この日時にフォームを自動クローズ。例：開催前日12:00", "system")
ws1.cell(row=1, column=18).comment = Comment("未実施 / 進行中 / 完了 / スキップ", "system")

# Example rows from real data
examples = [
    ["E001", 19, "グルコン",        "2/26グルコン",         "2026/2/26", '=TEXT(E2,"aaa")', "10:00", "11:00", "ヴォンドラ髙橋若菜", "19期", "https://us02web.zoom.us/j/83125213095", "831 2521 3095", "https://ws.formzu.net/dist/S16631693/", "2026/2/23", "2026/2/25 12:00", "", "", "完了"],
    ["E002", 19, "グルコン",        "3/12グルコン",         "2026/3/12", '=TEXT(E3,"aaa")', "21:00", "22:00", "ヴォンドラ髙橋若菜", "19期", "https://us02web.zoom.us/j/83125213095", "831 2521 3095", "https://ws.formzu.net/dist/S16631693/", "2026/3/9",  "2026/3/11 12:00", "", "", "完了"],
    ["E003", 19, "ビギナーグルコン","3/13ビギナーグルコン", "2026/3/13", '=TEXT(E4,"aaa")', "21:00", "22:00", "神保麻紀先生",       "19期", "",                                        "",              "https://ws.formzu.net/dist/S16942447/", "2026/3/10", "2026/3/12 12:00", "", "", "完了"],
    ["E004", 19, "課題作業会",      "3/31課題作業会",       "2026/3/31", '=TEXT(E5,"aaa")', "21:00", "22:00", "山下直子先生",        "19期", "https://us06web.zoom.us/j/82579711397", "825 7971 1397", "https://ws.formzu.net/fgen/S99301863/",  "2026/3/28", "2026/3/30 12:00", "", "", "完了"],
    ["E005", 19, "講義配信",        "第1回講義(ブランディング)","2026/2/16",'=TEXT(E6,"aaa")', "07:00", "",     "",                    "19期", "",                                        "",              "",                                       "",           "",              "https://vimeo.com/xxxxxxx", "第1回：ブランディング", "完了"],
    # blank template rows for user
]
for i, row_vals in enumerate(examples, 2):
    for c, v in enumerate(row_vals, 1):
        cell = ws1.cell(row=i, column=c, value=v)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        # yellow for input columns; gray for auto (曜日)
        if c == 6:  # 曜日 auto
            cell.fill = REF_FILL
        else:
            cell.fill = INPUT_FILL

# add blank input rows
for i in range(len(examples)+2, len(examples)+22):
    for c in range(1, len(headers1)+1):
        cell = ws1.cell(row=i, column=c)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        if c == 6:
            cell.value = f'=IF(E{i}="","",TEXT(E{i},"aaa"))'
            cell.fill = REF_FILL
        else:
            cell.fill = INPUT_FILL

# =================================================================
# Sheet 2: 共通設定
# =================================================================
ws2 = wb.create_sheet("② 共通設定")
set_widths(ws2, [4, 30, 70, 40])
ws2["B1"] = "キー"
ws2["C1"] = "値"
ws2["D1"] = "備考"
style_header(ws2, 1, 4)
ws2.cell(row=1, column=1).fill = HEADER_FILL
ws2.freeze_panes = "B2"

settings = [
    ("会員サイトURL",         "https://shinealight-xpvyol.club.hotmart.com/login", "Hotmartログイン"),
    ("会員サイトアクセス方法", "https://vimeo.com/1161412964/8fe81ba1d6?share=copy&fl=sv&fe=ci", "パスワード再設定動画"),
    ("19期FBグループURL",      "https://www.facebook.com/groups/1885679962314787", ""),
    ("19期継続FBグループURL",  "https://www.facebook.com/groups/841183151841115", ""),
    ("課題チェックシートフォルダ","https://drive.google.com/drive/folders/12F9Xrsb9IrE7CbrxKGVLv0MNjXRF3Gj0?usp=sharing", ""),
    ("事務局差出人名",         "Shine A Light 運営事務局", ""),
    ("事務局アドレス",         "your-office@example.com", "🟡 実際のアドレスに書き換える"),
    ("Chatwork APIトークン",   "", "🟡 Chatwork設定 > API から取得"),
    ("Chatwork ルームID(事務局)", "", "🟡 通知投稿先ルームのID"),
    ("Chatwork メンション先",  "[To:xxxxxx]", "🟡 自分のChatwork ID"),
    ("グルコンフォームURL",    "https://ws.formzu.net/dist/S16631693/", "ヴォンドラ先生グルコン用"),
    ("ビギナーフォームURL",    "https://ws.formzu.net/dist/S16942447/", "ビギナーグルコン用"),
    ("作業会フォームURL",      "https://ws.formzu.net/fgen/S99301863/", "課題作業会用"),
    ("下書き作成タイミング",   "07:00", "毎朝この時刻にGASが動く"),
]
for i, (k, v, note) in enumerate(settings, 2):
    ws2.cell(row=i, column=2, value=k).font = Font(name="Arial", bold=True, size=10)
    c = ws2.cell(row=i, column=3, value=v)
    c.font = BODY_FONT; c.alignment = WRAP; c.fill = INPUT_FILL; c.border = BORDER
    n = ws2.cell(row=i, column=4, value=note)
    n.font = NOTE_FONT; n.alignment = WRAP; n.border = BORDER
    ws2.cell(row=i, column=2).border = BORDER

# 署名 (multi-line)
ws2.cell(row=len(settings)+3, column=2, value="共通署名（本文末尾に自動付与）").font = Font(name="Arial", bold=True)
signature = ("---\n"
             "Shine A Light 運営事務局\n"
             "Mail: your-office@example.com")
c = ws2.cell(row=len(settings)+3, column=3, value=signature)
c.font = BODY_FONT; c.alignment = WRAP; c.fill = INPUT_FILL; c.border = BORDER
ws2.row_dimensions[len(settings)+3].height = 60

# 対象別 冒頭文
r = len(settings) + 5
ws2.cell(row=r, column=2, value="対象別 冒頭文（{{冒頭}} で差し込み）").font = Font(name="Arial", bold=True)
r += 1
opening_templates = [
    ("19期",                          "※このメールは、『19期\"Shine A Light\"』の皆様にお送りしております。"),
    ("19期継続",                     "※このメールは、『19期\"Shine A Light\"継続生』の皆様にお送りしております。"),
    ("19期継続(作業会特典対象者)",   "※このメールは、『19期継続\"Shine A Light\"』の皆様にお送りしております。"),
    ("18期から19期へ継続する受講生","※こちらのメールは18期から19期へ継続してくださっている受講生の皆様へお送りしております。"),
]
for k, v in opening_templates:
    ws2.cell(row=r, column=2, value=k).font = Font(name="Arial", size=10)
    c = ws2.cell(row=r, column=3, value=v)
    c.font = BODY_FONT; c.alignment = WRAP; c.fill = INPUT_FILL; c.border = BORDER
    ws2.cell(row=r, column=2).border = BORDER
    r += 1

# =================================================================
# Sheet 3: 本文テンプレ
# =================================================================
ws3 = wb.create_sheet("③ 本文テンプレ")
headers3 = ["テンプレID", "対応イベント種別", "タイミング", "送信日オフセット", "送信時刻", "件名テンプレ", "本文テンプレ", "備考"]
widths3 = [10, 20, 16, 14, 10, 55, 85, 25]
set_widths(ws3, widths3)
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, len(headers3))
ws3.freeze_panes = "A2"
ws3.row_dimensions[1].height = 32

# common body pieces
GRULCON_ADDENDUM = """【グルコンで添削をご希望の場合の注意点】

グルコンで「添削」を希望される方は、下記の要領でご提出ください。
円滑かつ公平な進行のため、ご協力をお願いいたします。

■ 対象物
・スライド ・SNS（プロフィール含む）・LP ・ホームページ ・Googleドキュメント など

■ 提出方法
1. 事前質問フォームに、対象物のURLリンクを必ず貼り付けてください。
2. リンクの閲覧権限を「リンクを知っている全員が閲覧可」に設定してください。

※グルコン直前の個別メッセージでのリンク送付は受け付けておりません。

■ 質問の書き方
「見てほしいです／添削してください」の一言ではなく、以下を具体的にお書きください。
・特に見てほしい箇所 ・現状と目的 ・相談内容・質問

■ 受付できないケース
・リンク未記載のご提出 ・個別メッセージでの直前送付 ・閲覧権限が付与されていないリンク"""

CLOSING = """ご案内は以上です。

ご不明点のお問い合わせは
当メールの返信にてご連絡ください。

{{署名}}"""

# ---- Templates ----
templates = [
    # グルコン
    ("T-G-01", "グルコン", "質問募集", "開催日 - 3", "07:00",
     "{{開催月}}/{{開催日D}}（{{曜日}}）{{開始時刻H}}時~開催グルコンの質問をお送りください",
     "{{冒頭}}\n\nこんにちは。\n\"Shine A Light\"運営事務局です。\n\n{{開催月}}月{{開催日D}}日（{{曜日}}）{{開始時刻H}}時〜開催のグルコンについてのご連絡です。\n\n課題をこなすのに、前へ進むのに、止まってしまっていませんか？\n\nグルコンへの質問を受け付けますので、ぜひ困っていること、相談したいことがありましたら、何でも、小さなことでもいいのでお送りください。\n\n【グルコン開催前の事前アンケートのお願い】\n\nグルコン当日に{{担当講師}}に相談したいこと&質問したいことを入力して、以下のアンケートフォームから送信してください。\n\n【回答期限：{{フォーム締切表記}}】までグルコンへの質問&相談をお受けしております。\n\n※グルコンの欠席連絡は不要です。質問をされる場合のみ、フォームに一言欠席の旨お書きください。\n\n＝＝＝＝＝＝\n\n▼グルコンの事前アンケート▼\n{{事前フォームURL}}\n\n＝＝＝＝＝＝\n\n" + GRULCON_ADDENDUM + "\n\n" + CLOSING,
     "開催3日前 朝"),

    ("T-G-02", "グルコン", "前日", "開催日 - 1", "07:00",
     "【重要】明日{{開始時刻H}}時〜グルコン詳細・事前質問受付は本日12時まで！",
     "{{冒頭}}\n\nこんにちは。\n\"Shine A Light\"運営事務局です。\n\n明日のグルコンの詳細のお知らせと事前質問アンケート送信のお願いです。\n\n明日【{{開催月}}月{{開催日D}}日（{{曜日}}）{{開始時刻H}}時〜{{終了時刻H}}時】にグルコンがございます。\n\n＝＝＝＝＝＝\n\n【グルコン参加URL（Zoom）】\n{{Zoomリンク}}\n\nミーティング ID: {{ミーティングID}}\n\n＝＝＝＝＝＝\n\n※グルコン動画は、後日アーカイブを送付いたします。\n※グルコンの欠席連絡は不要です。\n\n【グルコン事前質問受付は、回答期限は本日{{今日日付表記}}12時まで！】\n\n▼グルコンアンケート受付フォームURL▼\n{{事前フォームURL}}\n\n" + GRULCON_ADDENDUM + "\n\n" + CLOSING,
     "開催前日 朝"),

    ("T-G-03", "グルコン", "当日", "開催日 + 0", "07:00",
     "【重要】本日{{開始時刻H}}時〜グルコンの詳細です！",
     "{{冒頭}}\n\nこんにちは。\n\"Shine A Light\"運営事務局です。\n\n本日 【{{開催月}}月{{開催日D}}日（{{曜日}}）{{開始時刻H}}時〜{{終了時刻H}}時】にグルコンがございます。\n\n＝＝＝＝＝＝\n\n【グルコン参加URL（Zoom）】\n{{Zoomリンク}}\n\nミーティング ID: {{ミーティングID}}\n\n＝＝＝＝＝＝\n\n※グルコン動画は、後日アーカイブを送付いたします。\n本日ご都合がつかない方は、アーカイブをご活用ください。\n\n" + CLOSING.replace("{{署名}}", "それでは後ほど、よろしくお願いいたします。\n\n{{署名}}"),
     "開催当日 朝"),

    ("T-G-04", "グルコン", "アーカイブ", "開催日 + 1", "07:00",
     "【グルコン】{{開催月}}月{{開催日D}}日開催のアーカイブ動画です！",
     "{{冒頭}}\n\nこんにちは。\n\"Shine A Light\"運営事務局です。\n\n昨日のグルコンのアーカイブを会員サイトにアップしました！\n\n参加できなかった方や復習にご活用くださいませ。\n\n＝＝＝＝＝＝\n\n【会員サイト】＞＞\n{{会員サイトURL}}\n\n会員サイトへのアクセス方法＞＞\n{{会員サイトアクセス方法}}\n\n＝＝＝＝＝＝\n\n" + CLOSING,
     "開催翌日 朝"),

    # ビギナーグルコン
    ("T-B-01", "ビギナーグルコン", "質問募集", "開催日 - 3", "07:00",
     "【重要】{{開催月}}/{{開催日D}}（{{曜日}}）{{開始時刻H}}時~開催 ビギナーグルコンの質問をお送りください",
     "{{冒頭}}\n\nこんにちは。\n\"Shine A Light\"運営事務局です。\n\n{{開催月}}月{{開催日D}}日（{{曜日}}）{{開始時刻H}}時〜{{終了時刻H}}時にサポート講師によるビギナーグルコンを開催します。\n\n今回のサポート講師は{{担当講師}}です。\n\nグルコンへの質問を受け付けますので、ぜひ困っていること、相談したいことがありましたら、何でも、小さなことでもいいのでお送りください。\n\n【グルコン開催前の事前アンケートのお願い】\n\n【回答期限：{{フォーム締切表記}}】までグルコンへの質問&相談をお受けしております。\n\n▼グルコンの事前アンケート▼\n{{事前フォームURL}}\n\n" + CLOSING,
     "開催3日前 朝"),

    # 課題作業会
    ("T-W-01", "課題作業会", "質問募集", "開催日 - 3", "07:00",
     "【重要】{{開催月}}/{{開催日D}}（{{曜日}}）{{開始時刻H}}時~開催 課題作業会の質問をお送りください",
     "{{冒頭}}\n\nこんにちは。\n\"Shine A Light\"運営事務局です。\n\n{{開催月}}月{{開催日D}}日（{{曜日}}）{{開始時刻H}}:00〜開催の課題作業会についてのご連絡です。\n\n「作業会でこんな作業を一緒にしてほしい！」「〇〇に困っている！」ということがありましたら、事前にお伺いしますので事前フォームに入力ください。\n\n【回答期限：{{フォーム締切表記}}】まで作業会への質問&相談をお受けしております。\n\n＝＝＝＝＝＝\n\n▼作業会の事前フォーム▼\n{{事前フォームURL}}\n\n＝＝＝＝＝＝\n\n《作業会の詳細》\n【日時】{{開催月}}月{{開催日D}}日（{{曜日}}）{{開始時刻H}}時〜{{終了時刻H}}時\n・自由参加 ・途中参加、途中退出OK\n【担当講師】{{担当講師}}\n\n" + CLOSING,
     "開催3日前 朝"),

    ("T-W-02", "課題作業会", "前日", "開催日 - 1", "07:00",
     "【重要】明日{{開始時刻H}}時〜作業会の詳細・事前質問受付は本日12時まで！",
     "{{冒頭}}\n\nこんにちは。\n\"Shine A Light\"運営事務局です。\n\n明日の作業会の詳細のお知らせと事前フォーム送信のお願いです。\n\n明日【{{開催月}}月{{開催日D}}日（{{曜日}}）{{開始時刻H}}時〜{{終了時刻H}}時】に作業会がございます。\n\n＝＝＝＝＝＝\n\n【参加URL（Zoom）】\n{{Zoomリンク}}\n\nミーティング ID: {{ミーティングID}}\n\n＝＝＝＝＝＝\n\n【日時】{{開催月}}月{{開催日D}}日（{{曜日}}）{{開始時刻H}}時〜{{終了時刻H}}時\n【担当講師】{{担当講師}}\n\n【作業会の事前質問の回答期限は本日{{今日日付表記}}12時まで！】\n\n▼作業会の事前フォーム▼\n{{事前フォームURL}}\n\n" + CLOSING,
     "開催前日 朝"),

    ("T-W-03", "課題作業会", "当日", "開催日 + 0", "07:00",
     "【重要】本日{{開始時刻H}}時〜作業会の詳細です！",
     "{{冒頭}}\n\nこんにちは。\n\"Shine A Light\"運営事務局です。\n\n本日【{{開催月}}月{{開催日D}}日（{{曜日}}）{{開始時刻H}}時〜{{終了時刻H}}時】に作業会がございます。\n\n＝＝＝＝＝＝\n\n【参加URL（Zoom）】\n{{Zoomリンク}}\n\nミーティング ID: {{ミーティングID}}\n\n＝＝＝＝＝＝\n\n《作業会の詳細》\n【日時】{{開催月}}月{{開催日D}}日（{{曜日}}）{{開始時刻H}}時〜{{終了時刻H}}時\n【担当講師】{{担当講師}}\n\n" + CLOSING.replace("{{署名}}", "それでは後ほど、よろしくお願いいたします。\n\n{{署名}}"),
     "開催当日 朝"),

    ("T-W-04", "課題作業会", "アーカイブ", "開催日 + 1", "07:00",
     "【作業会】{{開催月}}月{{開催日D}}日開催のアーカイブ動画です！",
     "{{冒頭}}\n\nこんにちは。\n\"Shine A Light\"運営事務局です。\n\n昨日の作業会のアーカイブを会員サイトにアップしました！\n\n参加できなかった方や復習にご活用くださいませ。\n\n＝＝＝＝＝＝\n\n【会員サイト】＞＞\n{{会員サイトURL}}\n\n＝＝＝＝＝＝\n\n" + CLOSING,
     "開催翌日 朝"),

    # 講義配信
    ("T-L-01", "講義配信", "講義配信", "開催日 + 0", "07:00",
     "【重要】{{イベント名}}をお届けいたします",
     "{{冒頭}}\n\nこんにちは\nShine A Light 運営事務局です。\n\n本日、{{イベント名}}を会員サイトにアップいたしました。\n\n＝＝＝＝＝＝＝\n\n【会員サイト】＞＞\n{{会員サイトURL}}\n\n会員サイトへのアクセス方法＞＞\n{{会員サイトアクセス方法}}\n\n＝＝＝＝＝＝＝\n\n{{備考}}\n\n" + CLOSING,
     "配信当日 朝"),

    # 締切系
    ("T-D-01", "締切系", "受付開始", "受付開始日 + 0", "07:00",
     "【受付開始】{{イベント名}}",
     "{{冒頭}}\n\n（本文はイベントごとに個別編集する想定：受付開始のお知らせ）\n\n" + CLOSING,
     "受付開始日 朝"),

    ("T-D-02", "締切系", "締切前日", "締切日 - 1", "07:00",
     "【明日締切】{{イベント名}}",
     "{{冒頭}}\n\n明日 {{締切日表記}} が {{イベント名}} の締切です。\n\n" + CLOSING,
     "締切前日"),

    ("T-D-03", "締切系", "締切当日", "締切日 + 0", "07:00",
     "【本日締切】{{イベント名}}",
     "{{冒頭}}\n\n本日 {{締切日表記}} が {{イベント名}} の締切です。\n\n" + CLOSING,
     "締切当日"),
]

for i, t in enumerate(templates, 2):
    for c, v in enumerate(t, 1):
        cell = ws3.cell(row=i, column=c, value=v)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        cell.fill = INPUT_FILL
    ws3.row_dimensions[i].height = 240

# =================================================================
# Sheet 4: 差込項目 リファレンス
# =================================================================
ws4 = wb.create_sheet("④ 差込項目リファレンス")
set_widths(ws4, [22, 20, 55])
headers4 = ["差込タグ", "取得元", "例 / 説明"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, len(headers4))
ws4.freeze_panes = "A2"

variables = [
    ("{{期}}",                  "イベント日程シート", "19"),
    ("{{イベント名}}",           "イベント日程シート", "2/26グルコン"),
    ("{{イベント種別}}",         "イベント日程シート", "グルコン / ビギナーグルコン / 課題作業会 / 講義配信 …"),
    ("{{開催日}}",              "イベント日程シート", "2026/2/26"),
    ("{{開催月}}",              "自動（開催日から）",  "2"),
    ("{{開催日D}}",             "自動（開催日から）",  "26"),
    ("{{曜日}}",                "自動（開催日から）",  "木"),
    ("{{開始時刻H}}",           "自動（開始時刻から）", "10"),
    ("{{終了時刻H}}",           "自動（終了時刻から）", "11"),
    ("{{担当講師}}",             "イベント日程シート",  "ヴォンドラ髙橋若菜 / 山下直子先生 …"),
    ("{{対象}}",                "イベント日程シート",  "19期 / 19期継続 など"),
    ("{{Zoomリンク}}",           "イベント日程シート",  "https://us02web.zoom.us/j/..."),
    ("{{ミーティングID}}",       "イベント日程シート",  "831 2521 3095"),
    ("{{事前フォームURL}}",      "イベント日程シート",  "https://ws.formzu.net/dist/..."),
    ("{{フォーム締切表記}}",     "自動（フォーム締切日時から）", "2月25日（水）12時"),
    ("{{今日日付表記}}",         "自動（実行日から）", "2月25日（水）"),
    ("{{アーカイブURL}}",        "イベント日程シート",  "https://vimeo.com/..."),
    ("{{備考}}",                "イベント日程シート",  "第1回：ブランディング など"),
    ("{{冒頭}}",                "共通設定（対象別）",  "※このメールは、『19期\"Shine A Light\"』の皆様にお送りしております。"),
    ("{{署名}}",                "共通設定",           "Shine A Light 運営事務局 …"),
    ("{{会員サイトURL}}",        "共通設定",           "https://shinealight-xpvyol.club.hotmart.com/login"),
    ("{{会員サイトアクセス方法}}","共通設定",           "https://vimeo.com/1161412964/..."),
    ("{{FBグループURL}}",        "共通設定",           "https://www.facebook.com/groups/..."),
]
for i, v in enumerate(variables, 2):
    for c, val in enumerate(v, 1):
        cell = ws4.cell(row=i, column=c, value=val)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        cell.fill = REF_FILL

# reorder tabs so 使い方 is first
wb._sheets = [ws0, ws1, ws2, ws3, ws4]
wb.save(OUT)
print("wrote", OUT)
