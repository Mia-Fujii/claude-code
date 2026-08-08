"""Rebuild the full 20期 workbook with:
- 新しい列構成: A:内容 | B:日程 | C:日程短 | D:開始時間 | E:終了時間 | F:担当者 | G:メールセット | H:メール3日前 | ...
- メールテンプレートも修正版(19期ベース)に差し替え
- 使い方 / 基本設定 は既存を保持
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from datetime import datetime, time

SRC_ENHANCED = "/home/user/claude-code/scratchpad/ShineALight_20ki_Enhanced.xlsx"
SRC_TEMPLATES = "/home/user/claude-code/scratchpad/MailTemplates_20ki_v2.xlsx"
OUT = "/home/user/claude-code/scratchpad/ShineALight_20ki_v3.xlsx"

# ---------- Styles ----------
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
INPUT_FILL  = PatternFill("solid", fgColor="FFF2CC")   # 🟡
NEW_FILL    = PatternFill("solid", fgColor="E2EFDA")   # 🟢
REF_FILL    = PatternFill("solid", fgColor="E7E6E6")   # 🩶 auto
BODY_FONT   = Font(name="Arial", size=10)
WRAP        = Alignment(wrap_text=True, vertical="top")
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORDER = Border(*[Side(style="thin", color="BFBFBF")]*4)


def parse_time_range(time_val):
    """Return (start_str, end_str) from a value like '10:00〜11:00', '07:00:00', or time obj."""
    if time_val is None or time_val == "":
        return "", ""
    if isinstance(time_val, time):
        h = time_val.hour
        m = time_val.minute
        return f"{h:02d}:{m:02d}", ""
    s = str(time_val).strip()
    # try to split by 〜 or ~
    for sep in ["〜", "~", "-", "ー"]:
        if sep in s:
            parts = s.split(sep, 1)
            return _clean_time(parts[0]), _clean_time(parts[1])
    # single time
    return _clean_time(s), ""

def _clean_time(t):
    t = str(t).strip()
    # e.g. "10:00:00" → "10:00"
    if t.count(":") >= 2:
        return ":".join(t.split(":")[:2])
    return t

# ---------- Load sources ----------
wb_src = openpyxl.load_workbook(SRC_ENHANCED)
wb_tpl = openpyxl.load_workbook(SRC_TEMPLATES)
tpl_ws = wb_tpl["メールテンプレート"]

# ---------- Build new workbook ----------
wb = openpyxl.Workbook()
wb.remove(wb.active)

# --- Sheet: 使い方 (copy from src) ---
src_ws = wb_src["使い方"]
dst_ws = wb.create_sheet("使い方")
for col_letter in ["A", "B", "C"]:
    dst_ws.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width or 20
for row in src_ws.iter_rows():
    for cell in row:
        new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
        if cell.has_style:
            new_cell.font = Font(name=cell.font.name, bold=cell.font.bold, italic=cell.font.italic,
                                 color=cell.font.color, size=cell.font.size)
            new_cell.alignment = Alignment(wrap_text=cell.alignment.wrap_text,
                                           vertical=cell.alignment.vertical,
                                           horizontal=cell.alignment.horizontal)
    if src_ws.row_dimensions[cell.row].height:
        dst_ws.row_dimensions[cell.row].height = src_ws.row_dimensions[cell.row].height

# --- Sheet: 基本設定 (copy from src, no changes) ---
src_ws = wb_src["基本設定"]
dst_ws = wb.create_sheet("基本設定")
for col_letter in ["A", "B", "C"]:
    if src_ws.column_dimensions[col_letter].width:
        dst_ws.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width
for row in src_ws.iter_rows():
    for cell in row:
        new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
        if cell.has_style:
            new_cell.font = Font(name=cell.font.name, bold=cell.font.bold, italic=cell.font.italic,
                                 color=cell.font.color, size=cell.font.size)
            new_cell.alignment = Alignment(wrap_text=cell.alignment.wrap_text,
                                           vertical=cell.alignment.vertical,
                                           horizontal=cell.alignment.horizontal)
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.value not in (None, '00000000'):
                new_cell.fill = PatternFill("solid", fgColor=cell.fill.fgColor.value)
            new_cell.border = BORDER

# --- Sheet: タスク管理 (NEW column layout) ---
src_ws = wb_src["タスク管理"]
dst_ws = wb.create_sheet("スケジュール")

# 新しい列構成 (会員サイト/事前フォームURL/ID を削除 → 16列)
NEW_HEADERS = [
    "内容", "日程", "日程短",           # 3
    "開始時間", "終了時間", "担当者",    # 3
    "メールセット", "メール3日前", "メール前日", "質問まとめ",  # 4
    "メール当日", "アーカイブ送付",       # 2
    "Zoomソース", "Zoomリンク",         # 2
    "ミーティングID", "ステータス",       # 2
]
NEW_WIDTHS = [22, 12, 10, 10, 10, 18, 12, 12, 12, 10, 12, 12, 20, 40, 16, 12]

for i, w in enumerate(NEW_WIDTHS, 1):
    dst_ws.column_dimensions[get_column_letter(i)].width = w

for i, h in enumerate(NEW_HEADERS, 1):
    c = dst_ws.cell(row=1, column=i, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER

dst_ws.row_dimensions[1].height = 32
dst_ws.freeze_panes = "A2"

# 旧列インデックス (Enhanced xlsx)
OLD = {
    "content": 1, "date": 2, "time": 3, "manager": 4,
    "dMailSet": 5, "d3Days": 6, "dDayBefore": 7, "dQuestion": 8,
    "dDayOf": 9, "dArchive": 10, "memberSite": 11,
    "id": 12, "zoomSource": 13, "zoomUrl": 14, "meetingId": 15,
    "formUrl": 16, "status": 17,
}

# 各行を新レイアウトに変換
new_row_idx = 2
for src_row_idx in range(2, src_ws.max_row + 1):
    content = src_ws.cell(row=src_row_idx, column=OLD["content"]).value
    if not content:
        continue

    old_time = src_ws.cell(row=src_row_idx, column=OLD["time"]).value
    start_time, end_time = parse_time_range(old_time)

    # write new row
    values = [
        content,                                                     # A 内容
        src_ws.cell(row=src_row_idx, column=OLD["date"]).value,      # B 日程
        f'=IF(B{new_row_idx}="","",TEXT(B{new_row_idx},"M/d（aaa）"))',  # C 日程短
        start_time,                                                   # D 開始時間
        end_time,                                                     # E 終了時間
        src_ws.cell(row=src_row_idx, column=OLD["manager"]).value,   # F 担当者
        src_ws.cell(row=src_row_idx, column=OLD["dMailSet"]).value,  # G
        src_ws.cell(row=src_row_idx, column=OLD["d3Days"]).value,    # H
        src_ws.cell(row=src_row_idx, column=OLD["dDayBefore"]).value,# I
        src_ws.cell(row=src_row_idx, column=OLD["dQuestion"]).value, # J
        src_ws.cell(row=src_row_idx, column=OLD["dDayOf"]).value,    # K
        # L アーカイブ送付: 動画配信以外は全部 新formula (D列=開始時間参照)
        ("-" if "動画配信" in str(content or "")
         else f'=IF(IFERROR(TIMEVALUE(D{new_row_idx}),D{new_row_idx})>TIME(13,0,0),B{new_row_idx}+1,B{new_row_idx})'),
        src_ws.cell(row=src_row_idx, column=OLD["zoomSource"]).value,# M
        src_ws.cell(row=src_row_idx, column=OLD["zoomUrl"]).value,   # N
        src_ws.cell(row=src_row_idx, column=OLD["meetingId"]).value, # O
        src_ws.cell(row=src_row_idx, column=OLD["status"]).value or "未実施",  # P
    ]

    for col_i, v in enumerate(values, 1):
        cell = dst_ws.cell(row=new_row_idx, column=col_i, value=v)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        # 色付け
        if col_i == 3:  # 日程短 (auto formula)
            cell.fill = REF_FILL
        elif col_i == 13:  # Zoomソース (auto-assigned)
            cell.fill = NEW_FILL
        else:
            cell.fill = INPUT_FILL

    new_row_idx += 1

# ヘッダのコメント
dst_ws.cell(row=1, column=3).comment = Comment("=IF(B2=\"\",\"\",TEXT(B2,\"M/d（aaa）\")) で自動計算", "system")
dst_ws.cell(row=1, column=13).comment = Comment(
    "選択肢：若菜グルコン共通 / 若菜マンデー共通 / 個別 / 未定(サポート講師待ち) / Zoomなし",
    "system"
)

# --- Sheet: メールテンプレート (from MailTemplates v2) ---
dst_ws = wb.create_sheet("メールテンプレート")
for col_letter, width in [("A", 26), ("B", 14), ("C", 55), ("D", 90), ("E", 20)]:
    dst_ws.column_dimensions[col_letter].width = width
for row in tpl_ws.iter_rows():
    for cell in row:
        new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
        if cell.has_style:
            new_cell.font = Font(name=cell.font.name, bold=cell.font.bold, italic=cell.font.italic,
                                 color=cell.font.color, size=cell.font.size)
            new_cell.alignment = Alignment(wrap_text=cell.alignment.wrap_text,
                                           vertical=cell.alignment.vertical,
                                           horizontal=cell.alignment.horizontal)
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.value not in (None, '00000000'):
                new_cell.fill = PatternFill("solid", fgColor=cell.fill.fgColor.value)
            new_cell.border = BORDER
    if tpl_ws.row_dimensions[cell.row].height:
        dst_ws.row_dimensions[cell.row].height = tpl_ws.row_dimensions[cell.row].height
dst_ws.freeze_panes = "A2"

wb.save(OUT)
print(f"wrote {OUT}")

# Verify
wb2 = openpyxl.load_workbook(OUT)
print(f"Sheets: {wb2.sheetnames}")
ws = wb2["スケジュール"]
print(f"スケジュール: {ws.max_row} rows x {ws.max_column} cols")
# Print first few rows to verify
for r in range(1, 4):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 7)]
    print(f"  Row {r}: {vals}")
